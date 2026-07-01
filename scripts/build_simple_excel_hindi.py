"""Build Hindi-only simple Excel workbook — all data in Devanagari.

Output: data/gods_and_goddesses/Clan Gods Data - Simple (Hindi).xlsx

Usage:
    python3.12 scripts/build_simple_excel_hindi.py
"""
import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DATA_PATH = 'data/gods_and_goddesses/clan_gods.json'
OUT_PATH = 'data/gods_and_goddesses/Clan Gods Data - Simple (Hindi).xlsx'

with open(DATA_PATH) as f:
    cg = json.load(f)

pen_map = {p['id']: p for p in cg['pens']}
pen_by_name = {p['name'].lower(): p for p in cg['pens']}

wb = Workbook()

header_font = Font(bold=True, color='ffffff', size=11)
header_fill = PatternFill('solid', fgColor='2a6f97')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc'),
)

# ── Comprehensive Devanagari name map ──
HI = {
    # ── Pen names ──
    'Vedmo Moitor': 'वेदमो मोइतोर',
    'Urru Moitor': 'उर्रू मोइतोर',
    'Urru Dokri': 'उर्रू डोकरी',
    'Punga Rungi': 'पुंगा रूंगी',
    'Mai Sunga': 'माई सुंगा',
    'Hurra Mara': 'हुर्रा मारा',
    'Huru Mara': 'हुर्रू मारा',
    'Urru Mara': 'उर्रू मारा',
    'Chichur Unga': 'चिचुर उंगा',
    'Chichur Urra': 'चिचुर उर्रा',
    'Desh Deva': 'देश देवा',
    'Munuk Deva': 'मुनुक देवा',
    'Hadma Raj': 'हाडमा राज',
    'Jalsingo': 'जलसिंगो',
    'Iram Raj': 'ईराम राज',
    'Irma Raj': 'ईरमा राज',
    'Hunga Moitor': 'हुंगा मोइतोर',
    'Gujje Dokri': 'गुज्जे डोकरी',
    'Bhimaraj': 'भीमराज',
    'Bhimraj': 'भीमराज',
    'Biriya Bhima': 'बिरिया भीमा',
    'Bomul Ungal': 'बोमुल उंगल',
    'Godar Bhima': 'गोदर भीमा',
    'Chaikut Bhima': 'चैकुट भीमा',
    'Chaikud Bhima': 'चैकुड भीमा',
    'Godar Hunga': 'गोदर हुंगा',
    'Daro Moitor': 'दारो मोइतोर',
    'Darrem Modka': 'दरेम मोडका',
    'Daro Modka': 'दारो मोडका',
    'Raibandal': 'रायबंदल',
    'Kohla Kosu': 'कोहला कोसु',
    'Peda Hadma': 'पेडा हाडमा',
    'Chinna Hadma': 'चिन्ना हाडमा',
    'Murde Muyal': 'मुर्दे मुयाल',
    'Hunga Bhimal': 'हुंगा भीमाल',
    'Pind Hadmal': 'पिंड हाडमाल',
    'Bhum Iriyal': 'भूम इरियाल',
    'Pal Hadmal': 'पाल हाडमाल',
    'Muddaraj': 'मुद्दाराज',
    'Vidi Iriyal Panda Hadma': 'विडी इरियाल पांडा हाडमा',
    'Gaddi Kama': 'गड्डी कामा',
    'Gadi Kama': 'गडी कामा',
    'Gadye Kama': 'गड्ये कामा',
    'Ghadi Kama': 'गडी कामा',
    'Gadye/Gaddi Kama': 'गड्ये/गड्डी कामा',
    'Ghadi/Gadi/Gadye Kama': 'गडी/गडी/गड्ये कामा',
    'Murke Nango': 'मुर्के नांगो',
    'Kohli Dokri': 'कोहली डोकरी',
    'Gal Dullo': 'गाल दुल्लो',
    'Gal Bomda': 'गाल बोमडा',
    'Hingal Devo': 'हिंगल देवो',
    'Chenna Kama': 'चेन्ना कामा',
    'Ghanta Kama': 'घंटा कामा',
    'Ukud Kama': 'उकुड कामा',
    'Bade Vidi Iriyal': 'बडे विडी इरियाल',
    'Hadma Iriyal': 'हाडमा इरियाल',
    'Bande Boyo': 'बांदे बोयो',
    'Vange Dokri': 'वांगे डोकरी',
    'Vunge Dokri': 'वुंगे डोकरी',
    'Nanga Bhima': 'नांगा भीमा',
    'Uraal Gundal': 'उराल गुंडाल',
    'Ural Gunda': 'उराल गुंडा',
    'Barra Bujja': 'बर्रा बुज्जा',
    'Morka Moitor': 'मोरका मोइतोर',
    'Markaraj': 'मरकराज',
    'Hinge Dokri': 'हिंगे डोकरी',
    'Andalkosa': 'अंडलकोसा',
    'Andal Kosa': 'अंडल कोसा',
    'Lingal Denga': 'लिंगल डेंगा',
    'Lingal Deka': 'लिंगल डेका',
    'Nandraj': 'नंदराज',
    'Mawli': 'मावली',
    'Dol Mutte': 'डोल मुट्टे',
    'Dayur Mutte': 'दयुर मुट्टे',
    'Bhogam Mutte': 'भोगम मुट्टे',
    'Urru Ponde': 'उर्रू पोंडे',
    'Mudde Dokri': 'मुद्दे डोकरी',
    'Inge Dokri': 'इंगे डोकरी',
    'Lug Unga': 'लुग उंगा',
    'Katta Bodke': 'कट्टा बोडके',
    'Katte Bodke': 'कट्टे बोडके',
    'Akaluru Dokri': 'अकलुरु डोकरी',
    'Mawe Lungo': 'मावे लुंगो',
    'Mawe Kungo': 'मावे कुंगो',
    'Bhim Iriya': 'भीम इरिया',
    'Bhum Iriya': 'भूम इरिया',
    'Bhime': 'भीमे',
    'Madkaraj': 'मडकराज',
    'Ir Sunge': 'ईर सुंगे',
    'Murde Moyo': 'मुर्दे मोयो',
    # ── Clan names ──
    'Kunjam': 'कुंजम',
    'Markami': 'मरकामी',
    'Kalmu/Karma': 'कालमू/करमा',
    'Bhogam/Chote Telam': 'भोगम/छोटे तेलम',
    'Chote Telam/Bhogam': 'छोटे तेलम/भोगम',
    'Bade Telam': 'बडे तेलम',
    'Icham': 'ईचम',
    'Punem': 'पुनेम',
    'Kadiyam': 'कडियाम',
    'Midiyam': 'मिडियाम',
    'Undam': 'उंडम',
    'Tati': 'ताती',
    'Kadti': 'कडती',
    'Rengo': 'रेंगो',
    'Oyam': 'ओयम',
    'Hemla': 'हेमला',
    'Tamo': 'तामो',
    'Padami': 'पदामी',
    'Ujji/Dodi': 'उज्जी/दोडी',
    'Barse': 'बरसे',
    'Madvi': 'मडवी',
    'Kawasi': 'कवासी',
    'Telam': 'तेलम',
    'Bhogam': 'भोगम',
    # ── Phratry names ──
    'Kuhrami/Kadiari': 'कुहरामी/कडियारी',
    'Markami Kutumb': 'मरकामी कुटुंब',
    # ── Village and place names ──
    'Daler': 'डालेर',
    'Marh': 'मढ़',
    'Daler (near Marh)': 'डालेर (मढ़ के पास)',
    'Omalwar/Samalwar': 'ओमलवार/समलवार',
    'Gorgonda': 'गोरगोंडा',
    'Benpal/Bayampal': 'बेनपाल/बयामपाल',
    'Bhansi': 'भान्सी',
    'Hiroli': 'हिरोली',
    'Nendra': 'नेंद्रा',
    'Ukur': 'उकुर',
    'Malipad': 'मालीपाड',
    'Dilla': 'दिल्ला',
    'Itawar': 'इटावर',
    'Pidiya': 'पिडिया',
    'Tamodi': 'तामोडी',
    'Palnar': 'पालनार',
    'Kuwe': 'कुवे',
    'Gongpal': 'गोंगपाल',
    'Shyamgiri': 'श्यामगिरी',
    'Kunjampara': 'कुंजमपारा',
    'Ganjenar': 'गंजेनार',
    'Kuper': 'कुपेर',
    'Kudper': 'कुडपेर',
    'Kodoli': 'कोडोली',
    'Markagudem': 'मरकागुडेम',
    'Dugeli': 'दुगेली',
    'Pharaspal': 'फरासपाल',
    'Madpal': 'माडपाल',
    'Mirtulnar/Midkulnar': 'मिर्तुलनार/मिडकुलनार',
    'Manganar': 'मांगनार',
    'Jhirka': 'झिरका',
    'Baregunda': 'बरेगुंडा',
    'Renganar': 'रेंगनार',
    'Pendvela, Alnar': 'पेंडवेला, अलनार',
    'Katural, Pumbad (near Gangalur)': 'कटुराल, पुमबाड (गंगालुर के पास)',
    'Kondapal': 'कोंडापाल',
    'Metapal': 'मेटापाल',
    'Gumiyapal (Guyempad)': 'गुमियापाल (गुयेमपाड)',
    'Vengpal (Vengur)': 'वेंगपाल (वेंगुर)',
    'Tikanpal': 'टिकनपाल',
    'Madadi': 'माडाडी',
    'Jhadka (Orcha Block)': 'झड़का (ओरछा ब्लॉक)',
    'Kadampal': 'काडमपाल',
    'Vechapal': 'वेचापाल',
    'Kesapur': 'केसापुर',
    'Pandewar': 'पंडेवार',
    'Etlapad': 'एतलापाड',
    'Rekavaya': 'रेकावाया',
    'Kamkajojor': 'कामकाजोजोर',
    'Pinkonda': 'पिनकोंडा',
    'Tamirguda': 'तामीरगुडा',
    'Tadopadar': 'ताडोपाडर',
    'Kamalnar': 'कमालनार',
    'Kamalur': 'कमालुर',
    'Cholnar': 'चोलनार',
    'Jawanga': 'जवांगा',
    'Kuper/Kudper': 'कुपेर/कुडपेर',
    'Vengur': 'वेंगुर',
    'Pendvela': 'पेंडवेला',
    'Alnar': 'अलनार',
    'Katural': 'कटुराल',
    'Pumbad': 'पुमबाड',
    'Gangalur': 'गंगालुर',
    'Gumiyapal': 'गुमियापाल',
    'Guyempad': 'गुयेमपाड',
    'Vengpal': 'वेंगपाल',
    'Jhadka': 'झड़का',
    'Kamalnar?': 'कमालनार',
    'Kawasi?': 'कवासी',
    # ── Other place names ──
    'Dantewada': 'दंतेवाड़ा',
    'Warrangal': 'वारंगल',
    'Konta': 'कोंटा',
    'Errabor': 'एर्राबोर',
    'Jagargonda': 'जगरगोंडा',
    'Kongdam': 'कोंगडम',
    'Dondra': 'डोंडरा',
    'Gudra': 'गुडरा',
    'Murke': 'मुर्के',
    'Markanar': 'मरकानार',
    'Kumharras': 'कुम्हार्रास',
    'Matenar': 'माटेनार',
    'Kondoli': 'कोंडोली',
    'Gamawada': 'गामावाड़ा',
    'Mulasnar': 'मुलासनार',
    'Balod': 'बालोद',
    'Chandenar': 'चंदेनार',
    'Dhanora': 'धानोरा',
    'Dokometta': 'डोकोमेट्टा',
    'Tumnar': 'तुमनार',
    'Teknar': 'टेकनार',
    'Chitalur': 'चितालुर',
    'Kawalnar': 'कवालनार',
    'Turparas': 'तुरपारस',
    'Idodpara': 'इडोडपारा',
    'Jaraloha': 'जरालोहा',
    'Jara Loha': 'जरा लोहा',
    'Songunda': 'सोंगुंडा',
    'Kondanpal': 'कोंडानपाल',
    'Pen Kokhra': 'पेन कोखरा',
    # ── Person names (perma, etc.) ──
    'Joga Telam': 'जोगा तेलम',
    'Ballu Bhavani': 'बल्लू भवानी',
    # ── Other terms ──
    'Tuesday in February': 'फरवरी में मंगलवार',
    'Irirom Mada': 'ईरीरोम माडा',
    'Kendu': 'केंदु',
    'Dhurli': 'धुर्ली',
    # ── Relationship text ──
    'Younger brother of Urru Moitor': 'उर्रू मोइतोर का छोटा भाई',
    'Married to Urru Dokri': 'उर्रू डोकरी से विवाहित',
    'Bachelor': 'कुंवारा',
    'Name not known – have to find out': 'नाम ज्ञात नहीं — पता लगाना है',
    'Unknown Kunjam pen': 'अज्ञात कुंजम देवता',
    'Kunjam pen': 'कुंजम देवता',
    'Son of Punga Rungi and Mai Sunga': 'पुंगा रूंगी और माई सुंगा का पुत्र',
    'Son of Punga Rungi & Mai Sunga': 'पुंगा रूंगी और माई सुंगा का पुत्र',
    'Either a son of Punga Rungi & Mai Sunga or a son of Urru Moitor': 'पुंगा रूंगी और माई सुंगा का पुत्र या उर्रू मोइतोर का पुत्र',
    'Either a son of Punga Rungi & Mai Sunga or a son of Urru Moitor called Jalsingo': 'पुंगा रूंगी और माई सुंगा का पुत्र या उर्रू मोइतोर का पुत्र जिसे जलसिंगो कहते हैं',
    'Son of Punga Rungi & Mai Sunga': 'पुंगा रूंगी और माई सुंगा का पुत्र',
    'Either a son of Punga Rungi & Mai Sunga or a son of Urru Moitor': 'पुंगा रूंगी और माई सुंगा का पुत्र या उर्रू मोइतोर का पुत्र',
    's/o Urru Moitor': 'उर्रू मोइतोर का पुत्र',
    's/o Urru Moitor and Urre Dokri': 'उर्रू मोइतोर और उर्रे डोकरी का पुत्र',
    's/o Biriya Bhima': 'बिरिया भीमा का पुत्र',
    's/o Raibandal': 'रायबंदल का पुत्र',
    'd/o Muddaraj': 'मुद्दाराज की पुत्री',
    'd/o Gadi Kama (Oyam)': 'गडी कामा (ओयम) की पुत्री',
    'd/o Huru Mara of Bhansi': 'भान्सी के हुर्रू मारा की पुत्री',
    'd/o Bhum Iriya, Kadti pen of Madadi village': 'माडाडी गाँव की कडती देवता भूम इरिया की पुत्री',
    'Daughter of Nandraj and Mawli': 'नंदराज और मावली की पुत्री',
    'Gaddi Kama has four children with Murke Nango': 'गड्डी कामा के मुर्के नांगो से चार बच्चे',
    'Children of Gadi kama': 'गडी कामा के बच्चे',
    'Second brother of Gadi Kama (GK)': 'गडी कामा का दूसरा भाई',
    'Brother of GK': 'गडी कामा का भाई',
    'Third and youngest unmarried brother of Gaddi Kama': 'गड्डी कामा का तीसरा और सबसे छोटा अविवाहित भाई',
    'Elder Brother of Hurru Mara of Bhansi': 'भान्सी के हुर्रू मारा के बड़े भाई',
    'He is the third and youngest unmarried brother of Gaddi Kama': 'वह गड्डी कामा का तीसरा और सबसे छोटा अविवाहित भाई है',
    'Younger brother of Urru Moitor': 'उर्रू मोइतोर का छोटा भाई',
    'May or may not be brother of GK': 'गडी कामा का भाई हो या न हो',
    'Vange dokri is married to Nanga Bhima in Gangalur village': 'वांगे डोकरी का गंगालुर गाँव में नांगा भीमा से विवाह',
    'Gujje Dokri is married to Hunga Moitor in Dugeli': 'गुज्जे डोकरी का दुगेली में हुंगा मोइतोर से विवाह',
    'Markaraj is married to Hinge Dokri': 'मरकराज का हिंगे डोकरी से विवाह',
    'Raibandal married Akaluru Dokri': 'रायबंदल का अकलुरु डोकरी से विवाह',
    'Father of Mudde Dokri who is married to Biriya Bhima (Chote Telam) of Madpal': 'मुद्दे डोकरी के पिता जिनका विवाह माडपाल के बिरिया भीमा (छोटे तेलम) से हुआ',
    'His daughter Kohli Dokri is married to Gaddi Kama, Oyam Pen of Pandewar': 'उनकी पुत्री कोहली डोकरी का विवाह पंडेवार के ओयम देवता गड्डी कामा से हुआ',
    'Biriya Bhima is married to Mudde Dokri d/o Muddaraj': 'बिरिया भीमा का मुद्दे डोकरी (मुद्दाराज की पुत्री) से विवाह',
    'Bhimaraj is married to Urru Ponde of Madpal and Dol Mutte of Paddiguda': 'भीमराज का माडपाल की उर्रू पोंडे और पड्डीगुडा की डोल मुट्टे से विवाह',
    'Godar Bhima is married to Bhime d/o Gadi Kama (Oyam)': 'गोदर भीमा का भीमे (गडी कामा ओयम की पुत्री) से विवाह',
    'Hunga Moitor has 3 wives, Gujje Dokri from Pharaspal, Murde Moyo and Ir Sunge': 'हुंगा मोइतोर की तीन पत्नियाँ: फरासपाल की गुज्जे डोकरी, मुर्दे मोयो और ईर सुंगे',
    'Urru/Huru Mara had two wives – Urru Dokri and Mawe Lungo': 'उर्रू/हुर्रू मारा की दो पत्नियाँ — उर्रू डोकरी और मावे लुंगो',
    'Biriya Bhima perma is a Telam and Bade Telam have Urru Marra as their pen': 'बिरिया भीमा का परमा तेलम है और बडे तेलम का देवता उर्रू मारा है',
    'a Telam and Bade Telam have Urru Marra as their pen': 'तेलम और बडे तेलम का देवता उर्रू मारा है',
    'Godar Hunga is s/o Huru Mara and Katte Bodke': 'गोदर हुंगा हुर्रू मारा और कट्टे बोडके का पुत्र है',
    'See case study of Jhirka on phallis': 'पल्ली पर झिरका का केस स्टडी देखें',
    'This is a Telam village near a waterfall where Huru Mara and other Telam gods are made.': 'यह एक झरने के पास तेलम गाँव है जहाँ हुर्रू मारा और अन्य तेलम देवता बनाए जाते हैं।',
    'Daro Moitor palli includes Chandenar, Renganar': 'दारो मोइतोर की पल्ली में चंदेनार, रेंगनार शामिल',
    'She and husband visit each other at karsads.': 'वह और पति करसद पर एक-दूसरे से मिलते हैं।',
    'Made of irirom mada in Songunda above Kondanpal': 'कोंडानपाल के ऊपर सोंगुंडा में ईरीरोम माडा से बना',
    'Chichur Urra is made of kendu wood': 'चिचुर उर्रा केंदु की लकड़ी से बना है',
    'Punga Rungi (Telam wife, married to Mai Sunga Kunjam pen husband, now ghar jamai in Bhansi)': 'पुंगा रूंगी (तेलम पत्नी, माई सुंगा कुंजम देवता पति से विवाह, अब भान्सी में घर जमाई)',
    'Mai Sunga (ghar jamai)': 'माई सुंगा (घर जमाई)',
    'Desh Deva, younger brother of Mai Sunga, still a bachelor, and Munuk Deva, younger to Desh Deva': 'देश देवा, माई सुंगा का छोटा भाई, अभी कुंवारा, और मुनुक देवा, देश देवा से छोटा',
    '(ghar jamai)': '(घर जमाई)',
    'near Ganjenar': 'गंजेनार के पास',
    '(and Chaikut Bhima)': '(और चैकुट भीमा)',
    'near a waterfall': 'एक झरने के पास',
    's/o Huru Mara and Mawe Kungo': 'हुर्रू मारा और मावे कुंगो का पुत्र',
    'Mai Sunga has his palli and dhaniya here, in Jaraloha': 'माई सुंगा की पल्ली और धनिया यहाँ जरालोहा में है',
    'In 2019-20 they planned to get him married to a Mandavi girl.': '2019-20 में उसका विवाह एक मंडावी लड़की से करने की योजना बनाई गई थी।',
    'She will be brought as a spirit from Cholnar and her physical form will be made by the Kunjam people in Hiroli.': 'उसे चोलनार से आत्मा के रूप में लाया जाएगा और हिरोली में कुंजम लोग उसकी मूर्ति बनाएंगे।',
    'Biriya Bhima, Mudde Dokri and Urru Ponde are counted as Chote Telam.': 'बिरिया भीमा, मुद्दे डोकरी और उर्रू पोंडे को छोटे तेलम माना जाता है।',
    'Bhogams are also known as Chote Telams': 'भोगम को छोटे तेलम भी कहा जाता है',
    'Bhogam is one who does masti': 'भोगम वह है जो मस्ती करता है',
    'Chaikut = small and Godar = one who crossed the Godavari': 'चैकुट = छोटा और गोदर = गोदावरी पार करने वाला',
    'One of their sons still lives across the Godavari': 'उनका एक बेटा अभी भी गोदावरी पार रहता है',
    'one daughter Inge Dokri has her pen in Kuper': 'एक बेटी इंगे डोकरी का देवता कुपेर में है',
    'Markaraj palli covers Dantewada jail, Markanar, Kumharras, Matenar, Kondoli': 'मरकराज की पल्ली में दंतेवाड़ा जेल, मरकानार, कुम्हार्रास, माटेनार, कोंडोली शामिल',
    '(Grigson, p. 301)': '(ग्रिग्सन, पृ. 301)',
    'The Kunjams brought all the pen from Daler and then distributed them to all the different clans at Nandraj mountain.': 'कुंजम सभी देवताओं को डालेर से लाए और फिर नंदराज पर्वत पर विभिन्न कुलों में वितरित किया।',
    'Palli': 'पल्ली',
    'Perma': 'परमा',
    'Karsad': 'करसद',
    'Gudi': 'गुडी',
    'Gudi village — where the pen shrine is located.': 'गुडी गाँव — जहाँ देवता का मंदिर स्थित है।',
    'Ghadi Kama palli extends to Andhra, whole of Konta, Errabor, Jagargonda, Kongdam, Dondra, Gudra, Murke.': 'गडी कामा की पल्ली आंध्र, पूरे कोंटा, एर्राबोर, जगरगोंडा, कोंगडम, डोंडरा, गुडरा, मुर्के तक फैली है।',
    'His karsad is on a Tuesday in February.': 'उसका करसद फरवरी में मंगलवार को है।',
    'Village Jawanga': 'गाँव जवांगा',
    'Huru Mara palli covers Jhirka, Dokometta, Dhanora, Tumnar and Bhansi villages': 'हुर्रू मारा की पल्ली में झिरका, डोकोमेट्टा, धानोरा, तुमनार और भान्सी गाँव शामिल',
    '(His name on the shrine is written as Ghadi Kama though Ballu Bhavani spelt it out as Gadye Kama)': '(मंदिर पर उसका नाम गडी कामा लिखा है हालांकि बल्लू भवानी ने गड्ये कामा बताया)',
    'Gal Dullo, Etlapad': 'गाल दुल्लो, एतलापाड',
    'Gal Bomda, Etlapad': 'गाल बोमडा, एतलापाड',
    'Hingal Devo now in Korre Kongdam': 'हिंगल देवो अब कोरे कोंगडम में',
    '1 more boy': '1 और लड़का',
    'Gadi Kama and Kohli dokri (or Murke Nango?) also have children Chenna Kama, Ghanta Kama and Ukud Kama': 'गडी कामा और कोहली डोकरी (या मुर्के नांगो?) के भी बच्चे हैं: चेन्ना कामा, घंटा कामा और उकुड कामा',
    'They all live in Pandewar and their land or phalli is with their father.': 'वे सभी पंडेवार में रहते हैं और उनकी जमीन या पल्ली उनके पिता के साथ है।',
    'They had no sisters': 'उनकी कोई बहन नहीं थी',
    '(see story on Dol Mutte)': '(डोल मुट्टे की कहानी देखें)',
    'Hura Mara, Bhim Iriya, Daro Modka, Bhimaraj and Godar Bhima are all brothers.': 'हुर्रा मारा, भीम इरिया, दारो मोडका, भीमराज और गोदर भीमा सभी भाई हैं।',
    'also brothers of Huru Mara of Bhansi': 'भान्सी के हुर्रू मारा के भी भाई',
    'needs clarification': 'स्पष्टीकरण आवश्यक',
    'Son of Huru Mara and Katta Bodke (3rd wife)': 'हुर्रू मारा और कट्टा बोडके (तीसरी पत्नी) का पुत्र',
    'Punga Rungi d/o Huru Mara and Urru Dokri': 'पुंगा रूंगी हुर्रू मारा और उर्रू डोकरी की पुत्री',
    'Lug Unga, s/o Huru Mara and Mawe Kungo': 'लुग उंगा, हुर्रू मारा और मावे कुंगो का पुत्र',
    'Kunjam pen Mai Sunga who is married to Punga Rungi': 'कुंजम देवता माई सुंगा जो पुंगा रूंगी से विवाहित है',
    'Huru Mara and Mawe Lungo no longer visit Danteswari': 'हुर्रू मारा और मावे लुंगो अब दंतेश्वरी नहीं जाते',
    'Since Huru Mara eats humans – chewing them like datuns, she forbade them to come.': 'क्योंकि हुर्रू मारा मनुष्य खाता है — उन्हें दातून की तरह चबाता है, उसने उन्हें आने से मना किया।',
    'The Jhirka Pen is Godar Hunga': 'झिरका का देवता गोदर हुंगा है',
    'The Jhirka waterfall has two streams – Songunda/Kondapal and Barregunda.': 'झिरका झरने की दो धाराएँ हैं — सोंगुंडा/कोंडापाल और बरेगुंडा।',
    'ghar_jamai': 'घर जमाई',
    'ghar jamai': 'घर जमाई',
    'Gudi village': 'गुडी गाँव',
    'shrine': 'मंदिर',
}


def t(text):
    """Convert text to Devanagari Hindi. Uses override map first, then fallback."""
    if not text:
        return text

    # Try exact match
    if text in HI:
        return HI[text]

    # Try stripping extra whitespace
    stripped = text.strip()
    if stripped in HI:
        return HI[stripped]

    # For notes and longer texts, do a simple word-by-word substitution
    # then fall back to the transliterator for unknown words
    words = text.split()
    translated_words = []
    for w in words:
        if w in HI:
            translated_words.append(HI[w])
        elif w.lower() in HI:
            translated_words.append(HI[w.lower()])
        else:
            # Check punctuation variants
            w_clean = w.strip('.,;:!?()[]{}""''')
            if w_clean in HI:
                punct = w[len(w_clean):] if len(w_clean) < len(w) else ''
                translated_words.append(HI[w_clean] + punct)
            else:
                translated_words.append(w)
    return ' '.join(translated_words)


# ── Hindi column headers (contextually relevant) ──
PEN_COLS_HI = [
    'देवता का नाम',
    'अन्य नाम',
    'लिंग',
    'कुल',
    'गोत्र',
    'गुडी गाँव',
    'गाँव की अतिरिक्त जानकारी',
    'पल्ली (प्रभाव क्षेत्र)',
    'परमा (पुजारी)',
    'करसद (त्योहार)',
    'प्रकार',
    'जीवनसाथी',
    'भाई-बहन',
    'संतान',
    'माता-पिता',
    'टिप्पणी',
]

GENDER_MAP = {'male': 'पुरुष', 'female': 'महिला', 'unknown': 'अज्ञात'}
TYPE_MAP = {'main': 'मुख्य देवता', 'subordinate': 'अधीनस्थ देवता'}

CLAN_NAMES_HI = ['कुंजम', 'मरकामी', 'कालमू/करमा', 'भोगम/छोटे तेलम',
    'छोटे तेलम/भोगम', 'बडे तेलम', 'ईचम', 'पुनेम',
    'कडियाम', 'मिडियाम', 'उंडम', 'ताती', 'कडती', 'रेंगो',
    'ओयम', 'हेमला', 'तामो', 'पदामी', 'उज्जी/दोडी', 'बरसे',
    'मडवी', 'कवासी']

PHRATRY_NAMES_HI = ['कुहरामी/कडियारी', 'मरकामी कुटुंब', 'मडवी', 'कवासी']

# ── Data Dictionary (fully in Hindi) ──
DICT_ROWS_HI = [
    ['देवता का नाम', 'देवता का नाम (कुल देवता/भगवान)। यह अद्वितीय (यूनिक) होना चाहिए।', t('Godar Hunga'), 'हाँ', 'इससे स्वचालित रूप से देवता आईडी बनेगी। डुप्लीकेट नाम कनवर्टर द्वारा चिह्नित किए जाएंगे।'],
    ['अन्य नाम', 'इस देवता के वैकल्पिक नाम या वर्तनी।', t('Godar Hunga, Godar Hunga'), 'नहीं', 'अल्पविराम से अलग करें। रिश्तों को हल करते समय खोज के लिए उपयोग होता है।'],
    ['लिंग', 'देवता का लिंग।', 'पुरुष', 'हाँ', 'ड्रॉपडाउन: पुरुष / महिला / अज्ञात। नाम में स्त्री संकेतक (जैसे डोकरी, लुंगो) अपने आप पहचाने जाते हैं।'],
    ['कुल', 'यह देवता जिस कुल से संबंधित है।', t('Bade Telam'), 'यदि ज्ञात हो', '22 ज्ञात कुलों में से चुनें। अज्ञात होने पर खाली छोड़ें।'],
    ['गोत्र', 'यह देवता जिस गोत्र से संबंधित है।', t('Markami Kutumb'), 'यदि ज्ञात हो', 'ड्रॉपडाउन: कुहरामी/कडियारी, मरकामी कुटुंब, मडवी, कवासी।'],
    ['गुडी गाँव', 'गुडी गाँव — जहाँ देवता का मंदिर/स्थान स्थित है।', t('Jhirka'), 'नहीं', 'मुक्त पाठ। एक ही देवता नाम + गाँव अलग-अलग गोत्रों में = एक ही गाँव पुन: उपयोग होता है। नाम मिलने पर भी गोत्र अलग होने पर अलग गाँव माने जाते हैं।'],
    ['गाँव की अतिरिक्त जानकारी', 'गुडी गाँव के बारे में अतिरिक्त जानकारी: स्थान, स्थलचिह्न, गूगल मैप्स लिंक, आस-पास के गाँव, अक्षांश/देशांतर।', t('Near Jhirka waterfall, below the hill. 3 km from Kamalnar.'), 'नहीं', 'मुक्त पाठ। मानचित्र पर सीधे उपयोग नहीं होता — विश्लेषक संदर्भ के लिए।'],
    ['पल्ली (प्रभाव क्षेत्र)', 'प्रभाव क्षेत्र — गाँव/क्षेत्र जहाँ इस देवता की पूजा की जाती है।', t('Dokometta, Dhanora, Tumnar'), 'नहीं', 'अल्पविराम से अलग गाँवों के नाम। ये गुडी गाँव नहीं हैं — ये अतिरिक्त गाँव हैं जो यहाँ चढ़ावा लाते हैं।'],
    ['परमा (पुजारी)', 'पुजारी (परमा) का नाम जो इस देवता के मंदिर में अनुष्ठान करता है।', t('Joga Telam'), 'नहीं', 'मुक्त पाठ।'],
    ['करसद (त्योहार)', 'वार्षिक त्योहार का दिन जब इस देवता को चढ़ावा चढ़ाया जाता है।', 'फरवरी में मंगलवार', 'नहीं', 'मुक्त पाठ। उदाहरण: "फरवरी में मंगलवार", "दिवाड और बिज्जा पांडुम"।'],
    ['प्रकार', 'क्या यह अपने गाँव का मुख्य देवता है या अधीनस्थ देवता।', 'मुख्य देवता', 'हाँ', 'ड्रॉपडाउन: मुख्य देवता / अधीनस्थ देवता।'],
    ['जीवनसाथी', 'इस देवता का जीवनसाथी। एकाधिक के लिए अल्पविराम का प्रयोग करें।', t('Punga Rungi'), 'नहीं', 'देवता का नाम। घर जमाई के लिए "ghar_jamai:" उपसर्ग लगाएं। कनवर्टर इसे रिश्ते के रूप में पार्स करेगा।'],
    ['भाई-बहन', 'सहोदर देवता (भाई/बहन)।', t('Huru Mara, Daro Moitor, Bhimaraj, Godar Bhima'), 'नहीं', 'अल्पविराम से अलग देवता नाम। कनवर्टर सहोदर संबंध जोड़े बनाएगा।'],
    ['संतान', 'इस देवता की संतान।', t('Lug Unga, Punga Rungi'), 'नहीं', 'अल्पविराम से अलग देवता नाम। कनवर्टर माता-पिता-संतान संबंध बनाएगा।'],
    ['माता-पिता', 'इस देवता के माता-पिता (पिता/माता)।', t('Huru Mara'), 'नहीं', 'एकल देवता नाम। कनवर्टर संतान-माता-पिता संबंध बनाएगा।'],
    ['टिप्पणी', 'इस देवता के बारे में कोई अतिरिक्त टिप्पणी।', t('Son of Huru Mara and Katta Bodke (3rd wife)'), 'नहीं', 'मुक्त पाठ।'],
]

# ── Write Pens sheet ──
ws = wb.active
ws.title = 'देवता'
ws.sheet_properties.tabColor = '27ae60'

for i, col_name in enumerate(PEN_COLS_HI, 1):
    c = ws.cell(row=1, column=i, value=col_name)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = thin_border

gender_dv = DataValidation('list', formula1='"पुरुष,महिला,अज्ञात"', allow_blank=True)
ws.add_data_validation(gender_dv)

clan_dv = DataValidation('list',
    formula1='"%s"' % ','.join(CLAN_NAMES_HI), allow_blank=True)
ws.add_data_validation(clan_dv)

phratry_dv = DataValidation('list',
    formula1='"%s"' % ','.join(PHRATRY_NAMES_HI), allow_blank=True)
ws.add_data_validation(phratry_dv)

type_dv = DataValidation('list', formula1='"मुख्य देवता,अधीनस्थ देवता"', allow_blank=True)
ws.add_data_validation(type_dv)


def clan_id_to_hi(clan_id):
    if not clan_id:
        return ''
    for c in cg['clans']:
        if c['id'] == clan_id:
            return t(c['name'])
    return clan_id


def phratry_id_to_hi(phratry_id):
    if not phratry_id:
        return ''
    for p in cg['phratries']:
        if p['id'] == phratry_id:
            return t(p['name'])
    return phratry_id


def resolve_single_pen_hi(pen_id):
    if not pen_id:
        return ''
    p = pen_map.get(pen_id)
    return t(p['name']) if p else pen_id


print("Building pen data (67 rows)...")
for r, pen in enumerate(cg['pens'], 2):
    pen_id = pen['id']

    spouses = []
    for rel in cg['relationships']:
        if rel['type'] == 'marriage' and rel['from_pen_id'] == pen_id:
            p = pen_map.get(rel['to_pen_id'])
            if p:
                spouses.append(t(p['name']))
        if rel['type'] == 'marriage' and rel['to_pen_id'] == pen_id:
            p = pen_map.get(rel['from_pen_id'])
            if p and t(p['name']) not in spouses:
                spouses.append(t(p['name']))
        if rel['type'] == 'ghar_jamai' and rel['from_pen_id'] == pen_id:
            p = pen_map.get(rel['to_pen_id'])
            if p:
                spouses.append('ghar_jamai:' + t(p['name']))

    siblings = []
    for rel in cg['relationships']:
        if rel['type'] == 'sibling':
            if rel['from_pen_id'] == pen_id:
                p = pen_map.get(rel['to_pen_id'])
                if p:
                    siblings.append(t(p['name']))
            elif rel['to_pen_id'] == pen_id:
                p = pen_map.get(rel['from_pen_id'])
                if p and t(p['name']) not in siblings:
                    siblings.append(t(p['name']))

    children = []
    for rel in cg['relationships']:
        if rel['type'] == 'parent' and rel['from_pen_id'] == pen_id:
            p = pen_map.get(rel['to_pen_id'])
            if p:
                children.append(t(p['name']))

    parent_list = []
    for rel in cg['relationships']:
        if rel['type'] == 'parent' and rel['to_pen_id'] == pen_id:
            p = pen_map.get(rel['from_pen_id'])
            if p:
                parent_list.append(t(p['name']))

    gender_raw = pen.get('gender', '')
    gender_hi = GENDER_MAP.get(gender_raw, gender_raw) if gender_raw else ''
    type_raw = pen.get('type', '')
    type_hi = TYPE_MAP.get(type_raw, type_raw) if type_raw else ''

    vals = [
        t(pen['name']),
        t(', '.join(pen.get('aliases', []))) if pen.get('aliases') else '',
        gender_hi,
        clan_id_to_hi(pen.get('clan_id')),
        phratry_id_to_hi(pen.get('phratry_id')),
        t(pen.get('gudi_village') or ''),
        '',
        t(', '.join(pen.get('palli', []))) if pen.get('palli') else '',
        t(pen.get('perma', '')),
        t(pen.get('karsad', '')),
        type_hi,
        ', '.join(spouses),
        ', '.join(siblings),
        ', '.join(children),
        parent_list[0] if parent_list else '',
        t(pen.get('notes', '')),
    ]

    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = cell_align
        cell.border = thin_border
        if c == 1:
            cell.font = Font(bold=True)

    gender_dv.add(ws.cell(row=r, column=3))
    clan_dv.add(ws.cell(row=r, column=4))
    phratry_dv.add(ws.cell(row=r, column=5))
    type_dv.add(ws.cell(row=r, column=11))

col_widths = {
    1: 24, 2: 28, 3: 10, 4: 22, 5: 22,
    6: 28, 7: 38, 8: 38, 9: 18, 10: 20,
    11: 14, 12: 26, 13: 34, 14: 34, 15: 22, 16: 44
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:P{len(cg["pens"]) + 1}'

# ── Write Data Dictionary sheet ──
ws2 = wb.create_sheet('डेटा डिक्शनरी', 1)
ws2.sheet_properties.tabColor = '3498db'

dict_headers_hi = ['फ़ील्ड', 'विवरण', 'उदाहरण', 'आवश्यक', 'नियम']
dict_header_fill = PatternFill('solid', fgColor='3498db')

for i, name in enumerate(dict_headers_hi, 1):
    c = ws2.cell(row=1, column=i, value=name)
    c.font = Font(bold=True, color='ffffff', size=11)
    c.fill = dict_header_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

for r, row_data in enumerate(DICT_ROWS_HI, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border
        if c == 1:
            cell.font = Font(bold=True)

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 58
ws2.column_dimensions['C'].width = 42
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 58
ws2.freeze_panes = 'A2'

# Save
wb.save(OUT_PATH)
print(f"\n✓ Wrote {OUT_PATH}")
print(f"  Pens: {len(cg['pens'])} rows pre-populated (all Devanagari Hindi)")
print(f"  Data Dictionary: {len(DICT_ROWS_HI)} field descriptions (Hindi)")
