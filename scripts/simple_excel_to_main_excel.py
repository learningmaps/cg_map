"""Convert Clan Gods Data - Simple.xlsx → Clan Gods Data.xlsx (main workbook).

Usage:
    python3 scripts/simple_excel_to_main_excel.py
    python3 scripts/simple_excel_to_main_excel.py --validate
"""
import json, re, sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SIMPLE_XLSX = 'data/gods_and_goddesses/Clan Gods Data - Simple.xlsx'
OUT_XLSX = 'data/gods_and_goddesses/Clan Gods Data.xlsx'
CG_JSON = 'data/gods_and_goddesses/clan_gods.json'

# ── Helpers ──

def make_id(name):
    if not name: return None
    name = name.strip().replace('\n', ' ').replace('\r', ' ')
    n = re.sub(r'[^a-z0-9_\s]', '', name.lower())
    return re.sub(r'\s+', '_', n.strip())

def parse_list(val):
    if not val: return []
    return [v.strip() for v in str(val).split(',') if v.strip()]

def parse_v_details(val):
    """Parse village_details cell into a list of v_codes/strings.
    
    Handles comma-separated, newline-separated, and mixed formats.
    Skips entries that look like descriptive notes (no 6-digit code).
    """
    if not val: return []
    raw = str(val).strip()
    if not raw: return []
    # Split by comma, newline, or semicolon
    parts = re.split(r'[\n,;]+', raw)
    results = []
    for p in parts:
        p = p.strip()
        if not p: continue
        results.append(p)
    return results

# ── Load Simple Excel ──

wb_simple = load_workbook(SIMPLE_XLSX)
ws_pens = wb_simple['Pens']

raw_rows = []
for row in ws_pens.iter_rows(min_row=2, max_row=ws_pens.max_row, values_only=True):
    name = row[0]
    if not name or '═══' in str(name):
        continue
    raw_rows.append({
        'name': str(name).strip(),
        'aliases': parse_list(row[1]),
        'gender': str(row[2] or 'unknown').strip(),
        'clan': str(row[3] or '').strip(),
        'phratry': str(row[4] or '').strip(),
        'village': str(row[5] or '').strip(),
        'village_details': str(row[6] or '').strip(),
        'palli': parse_list(row[7]),
        'perma': str(row[8] or '').strip(),
        'karsad': str(row[9] or '').strip(),
        'type': str(row[10] or 'main').strip(),
        'spouse': parse_list(row[11]),
        'siblings': parse_list(row[12]),
        'children': parse_list(row[13]),
        'parent': parse_list(row[14]),
        'notes': str(row[15] or '').strip(),
    })

print(f"Simple Excel rows (pen-village pairs): {len(raw_rows)}")

# ── Group raw rows by pen name (same pen can appear in multiple villages) ──

pen_groups = {}
pen_name_order = []
for r in raw_rows:
    nk = r['name'].lower()
    if nk not in pen_groups:
        pen_groups[nk] = {'first': r, 'rows': []}
        pen_name_order.append(nk)
    pen_groups[nk]['rows'].append(r)

print(f"Unique pens: {len(pen_groups)}")

# ── Load existing JSON for reference data ──

with open(CG_JSON) as f:
    cg = json.load(f)

# ── Build Phratries & Clans from existing (these are reference, stable) ──

phratries = cg['phratries']
clans = cg['clans']
clan_name_to_id = {}
for c in clans:
    clan_name_to_id[c['name'].lower()] = c['id']
    for alt in c['name'].split('/'):
        clan_name_to_id[alt.strip().lower()] = c['id']
clan_name_to_id['telam'] = 'telam_bade'

phratry_name_to_id = {}
for p in phratries:
    phratry_name_to_id[p['name'].lower()] = p['id']

if 'sodi' not in phratry_name_to_id.values():
    phratries.append({'id': 'sodi', 'name': 'Sodi', 'source': None, 'clan_ids': ['sodi']})
    phratry_name_to_id['sodi'] = 'sodi'
if 'sodi' not in clan_name_to_id.values():
    clans.append({'id': 'sodi', 'name': 'Sodi', 'phratry_id': 'sodi', 'notes': None})
    clan_name_to_id['sodi'] = 'sodi'

print(f"Phratries: {len(phratries)}, Clans: {len(clans)}")

# ── Resolve clan/phratry IDs from names ──

def resolve_clan_id(name):
    if not name: return None
    return clan_name_to_id.get(name.lower())

def resolve_phratry_id(name):
    if not name: return None
    pid = phratry_name_to_id.get(name.lower())
    if pid: return pid
    for p in phratries:
        if name.lower() in p['name'].lower():
            return p['id']
    return None

# ── Build Pens with IDs (one per unique pen name) ──

used_ids = set()
pens = []
for nk in pen_name_order:
    group = pen_groups[nk]
    sp = group['first']
    rows = group['rows']

    pid = make_id(sp['name'])
    if not pid:
        print(f"  WARN: Could not generate ID for '{sp['name']}'")
        continue
    if pid in used_ids:
        suffix = 2
        while f'{pid}_{suffix}' in used_ids:
            suffix += 1
        pid = f'{pid}_{suffix}'
    used_ids.add(pid)

    cl_id = resolve_clan_id(sp['clan'])
    ph_id = resolve_phratry_id(sp['phratry'])

    # Collect all v_codes from all rows for this pen
    all_vcodes = []
    all_village_names = []
    for r in rows:
        names = parse_v_details(r['village'])
        details = parse_v_details(r['village_details'])
        # v_code is the last 6-digit entry in village_details (or the only one)
        for d in details:
            d_stripped = d.strip()
            # Check if it's a 6-digit v_code
            if re.match(r'^\d{6}$', d_stripped):
                if d_stripped not in all_vcodes:
                    all_vcodes.append(d_stripped)
            else:
                # Could be text like "seems to be in Potali Village"
                # For v_code extraction, also check if there's a 6-digit number embedded
                nums = re.findall(r'\b\d{6}\b', d_stripped)
                for n in nums:
                    if n not in all_vcodes:
                        all_vcodes.append(n)
        # Collect village names
        for n in names:
            n_s = n.strip()
            if n_s and n_s not in all_village_names:
                all_village_names.append(n_s)

    # First v_code/name is the "primary" (backward compat)
    first_vc = all_vcodes[0] if all_vcodes else None
    first_vname = all_village_names[0] if all_village_names else (rows[0].get('village', '') or None)

    pens.append({
        'id': pid,
        'name': sp['name'],
        'aliases': sp['aliases'],
        'clan_id': cl_id,
        'phratry_id': ph_id,
        'gender': sp['gender'],
        'village_ids': all_vcodes,
        'village_names': all_village_names,
        'gudi_village_id': first_vc,
        'gudi_village': first_vname,
        'palli': sp['palli'],
        'perma': sp['perma'],
        'karsad': sp['karsad'],
        'type': sp['type'],
        'notes': sp['notes'],
        'spouse': sp['spouse'],
        'siblings': sp['siblings'],
        'children': sp['children'],
        'parent': sp['parent'],
    })

print(f"Pens built: {len(pens)}")

# ── Build Pen name → ID lookup (for relationships) ──
name_to_pid = {}
for p in pens:
    name_to_pid[p['name'].lower()] = p['id']
    for a in p['aliases']:
        name_to_pid[a.strip().lower()] = p['id']

def find_pen_id(ref_name):
    ref = ref_name.strip().lower()
    if ref in name_to_pid:
        return name_to_pid[ref]
    clean = re.sub(r'\(.*?\)', '', ref).strip()
    if clean and clean != ref and clean in name_to_pid:
        return name_to_pid[clean]
    for pname, pid in name_to_pid.items():
        if ref in pname:
            return pid
        if pname in ref and re.search(r'(?:^|\s)' + re.escape(pname) + r'(?:\s|$)', ref):
            return pid
    return None

# ── Build Relationships ──

relationships = []
seen_rels = set()

for p in pens:
    pid = p['id']
    # Spouse → marriage
    for s in p.get('spouse', []):
        s_clean = re.sub(r'\(.*?\)', '', s).strip()
        sp_id = find_pen_id(s)
        if sp_id and sp_id != pid:
            key = ('marriage', pid, sp_id)
            if key not in seen_rels:
                seen_rels.add(key)
                relationships.append({
                    'type': 'marriage', 'from_pen_id': pid, 'to_pen_id': sp_id,
                    'details': f'{p["name"]} married to {s}', 'source': 'Simple Excel'
                })
        elif not sp_id and s_clean:
            sp_id = find_pen_id(s_clean)
            if sp_id and sp_id != pid:
                key = ('marriage', pid, sp_id)
                if key not in seen_rels:
                    seen_rels.add(key)
                    relationships.append({
                        'type': 'marriage', 'from_pen_id': pid, 'to_pen_id': sp_id,
                        'details': f'{p["name"]} married to {s}', 'source': 'Simple Excel'
                    })
    # Siblings → sibling
    for s in p.get('siblings', []):
        sib_id = find_pen_id(s)
        if sib_id and sib_id != pid:
            key = ('sibling', pid, sib_id)
            if key not in seen_rels:
                seen_rels.add(key)
                relationships.append({
                    'type': 'sibling', 'from_pen_id': pid, 'to_pen_id': sib_id,
                    'details': f'{p["name"]} sibling of {s}', 'source': 'Simple Excel'
                })
    # Children → parent
    for c in p.get('children', []):
        ch_id = find_pen_id(c)
        if ch_id and ch_id != pid:
            key = ('parent', pid, ch_id)
            if key not in seen_rels:
                seen_rels.add(key)
                relationships.append({
                    'type': 'parent', 'from_pen_id': pid, 'to_pen_id': ch_id,
                    'details': f'{p["name"]} parent of {c}', 'source': 'Simple Excel'
                })
    # Parent → child
    for par in p.get('parent', []):
        par_id = find_pen_id(par)
        if par_id and par_id != pid:
            key = ('parent', par_id, pid)
            if key not in seen_rels:
                seen_rels.add(key)
                relationships.append({
                    'type': 'parent', 'from_pen_id': par_id, 'to_pen_id': pid,
                    'details': f'{par} parent of {p["name"]}', 'source': 'Simple Excel'
                })

print(f"Relationships built: {len(relationships)}")

# ── Build Villages from ALL raw rows (one village entry per (phratry, village_name, v_code)) ──

village_entries = {}  # key → {name, phratry_id, clan_id, v_code, pen_ids}

for r in raw_rows:
    vname = (r['village'] or '').strip()
    if not vname:
        continue
    phid = resolve_phratry_id(r['phratry'])
    clid = resolve_clan_id(r['clan'])

    # Find the v_code for this row
    vdetails = (r['village_details'] or '').strip()
    vcode = None
    if vdetails:
        nums = re.findall(r'\b\d{6}\b', vdetails)
        if nums:
            vcode = nums[-1]  # last 6-digit number

    # Match this row to a pen
    pid = None
    for p in pens:
        if p['name'].lower() == r['name'].lower():
            pid = p['id']
            break

    # Group by (phratry_id, vname_lower)
    lkey = vname.lower()
    # Within a phratry, a village name should be unique
    key = (phid, lkey, vcode)
    if key not in village_entries:
        village_entries[key] = {
            'name': vname,
            'phratry_id': phid,
            'clan_id': clid,
            'v_code': vcode,
            'pen_ids': [],
            'main_pen_id': None,
            'sub_pen_ids': [],
        }
    if pid and pid not in village_entries[key]['pen_ids']:
        village_entries[key]['pen_ids'].append(pid)

    # Prefer first pen as main
    if village_entries[key]['main_pen_id'] is None and pid:
        village_entries[key]['main_pen_id'] = pid

# Build villages list, mark main/sub pens
villages = []
for key, ve in village_entries.items():
    pen_ids_in_village = ve['pen_ids']
    main_pen = pen_ids_in_village[0] if pen_ids_in_village else None
    sub_pens = pen_ids_in_village[1:] if pen_ids_in_village else []

    villages.append({
        'name': ve['name'],
        'aliases': [],
        'phratry_id': ve['phratry_id'],
        'clan_id': ve['clan_id'],
        'main_pen_id': main_pen,
        'subordinate_pen_ids': sub_pens,
        'notes': '',
        'population_info': '',
        'id': ve['v_code'],
    })

print(f"Villages built: {len(villages)}")

# ── Validation ──
issues = []
pen_ids = {p['id'] for p in pens}
clan_ids = {c['id'] for c in clans}
phratry_ids = {p['id'] for p in phratries}
village_ids = {v['id'] for v in villages if v.get('id')}

for i, v in enumerate(villages):
    if v['phratry_id'] and v['phratry_id'] not in phratry_ids:
        issues.append(f'Village[{i}] "{v["name"]}": unknown phratry "{v["phratry_id"]}"')
    if v['clan_id'] and v['clan_id'] not in clan_ids:
        issues.append(f'Village[{i}] "{v["name"]}": unknown clan "{v["clan_id"]}"')
    if v['main_pen_id'] and v['main_pen_id'] not in pen_ids:
        issues.append(f'Village[{i}] "{v["name"]}": unknown main_pen_id "{v["main_pen_id"]}"')
    for sub_id in v.get('subordinate_pen_ids', []):
        if sub_id and sub_id not in pen_ids:
            issues.append(f'Village[{i}] "{v["name"]}": unknown subordinate pen ID "{sub_id}"')

for i, p in enumerate(pens):
    for gvi in p.get('village_ids', []):
        if gvi and gvi not in village_ids:
            issues.append(f'Pen[{i}] "{p["name"]}": unknown village_id "{gvi}"')

for i, r in enumerate(relationships):
    if r['from_pen_id'] not in pen_ids:
        issues.append(f'Relation[{i}]: unknown from_pen_id "{r["from_pen_id"]}"')
    if r['to_pen_id'] not in pen_ids:
        issues.append(f'Relation[{i}]: unknown to_pen_id "{r["to_pen_id"]}"')

if issues:
    print(f'VALIDATION: {len(issues)} issue(s)')
    for iss in issues:
        print(f'  * {iss}')
    if '--validate' in sys.argv:
        sys.exit(1)
else:
    print('Validation: PASS ✓')

if '--validate' in sys.argv:
    sys.exit(0)

# ── Write Main Excel ──

wb = Workbook()
header_font = Font(bold=True, color='ffffff', size=11)
header_fill = PatternFill('solid', fgColor='2a6f97')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc'),
)

def write_header(ws, cols):
    for i, (name, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = header_font; c.fill = header_fill
        c.alignment = header_align; c.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = width

# Sheet 1: README
ws = wb.active
ws.title = 'README'
ws.cell(row=1, column=1, value='CLAN GODS DATA — Analyst Workbook').font = Font(bold=True, size=14)
ws.cell(row=3, column=1, value='Generated from Clan Gods Data - Simple.xlsx')
ws.cell(row=4, column=1, value=f'{len(pens)} pens, {len(villages)} villages, {len(relationships)} relationships')

# Sheet 2: Phratries
ws = wb.create_sheet('Phratries', 1)
write_header(ws, [('id', 20), ('name', 25), ('source', 30), ('clan_ids', 60)])
for i, p in enumerate(phratries, 2):
    ws.cell(row=i, column=1, value=p['id']).border = thin_border
    ws.cell(row=i, column=2, value=p['name']).border = thin_border
    ws.cell(row=i, column=3, value=p.get('source', '')).border = thin_border
    ws.cell(row=i, column=4, value=','.join(p.get('clan_ids', []))).border = thin_border

# Sheet 3: Clans
ws = wb.create_sheet('Clans', 2)
write_header(ws, [('id', 20), ('name', 25), ('phratry_id', 20), ('notes', 40)])
for i, c in enumerate(clans, 2):
    ws.cell(row=i, column=1, value=c['id']).border = thin_border
    ws.cell(row=i, column=2, value=c['name']).border = thin_border
    ws.cell(row=i, column=3, value=c.get('phratry_id', '')).border = thin_border
    ws.cell(row=i, column=4, value=c.get('notes', '')).border = thin_border

# Sheet 4: Villages
ws = wb.create_sheet('Villages', 3)
write_header(ws, [
    ('name', 28), ('aliases', 20), ('phratry_id', 18), ('clan_id', 18),
    ('main_pen_id', 22), ('main_pen_name', 25), ('subordinate_pen_ids', 30),
    ('notes', 30), ('population_info', 20), ('id', 12)
])
for i, v in enumerate(villages, 2):
    ws.cell(row=i, column=1, value=v['name']).border = thin_border
    ws.cell(row=i, column=2, value=','.join(v.get('aliases', []))).border = thin_border
    ws.cell(row=i, column=3, value=v.get('phratry_id', '')).border = thin_border
    ws.cell(row=i, column=4, value=v.get('clan_id', '')).border = thin_border
    ws.cell(row=i, column=5, value=v.get('main_pen_id', '')).border = thin_border
    ws.cell(row=i, column=6, value='').border = thin_border
    ws.cell(row=i, column=7, value=','.join(v.get('subordinate_pen_ids', []))).border = thin_border
    ws.cell(row=i, column=8, value=v.get('notes', '')).border = thin_border
    ws.cell(row=i, column=9, value=v.get('population_info', '')).border = thin_border
    ws.cell(row=i, column=10, value=v.get('id', '')).border = thin_border
ws.auto_filter.ref = f'A1:J{len(villages)+1}'
ws.freeze_panes = 'A2'

# Sheet 5: Pens
ws = wb.create_sheet('Pens', 4)
write_header(ws, [
    ('id', 25), ('name', 28), ('aliases', 25), ('clan_id', 18), ('phratry_id', 18),
    ('gender', 10), ('village_ids', 20), ('gudi_village_id', 16), ('gudi_village_name', 28),
    ('palli', 40), ('perma', 18), ('karsad', 18), ('type', 12), ('notes', 50)
])
for i, p in enumerate(pens, 2):
    ws.cell(row=i, column=1, value=p['id']).border = thin_border
    ws.cell(row=i, column=2, value=p['name']).border = thin_border
    ws.cell(row=i, column=3, value=','.join(p.get('aliases', []))).border = thin_border
    ws.cell(row=i, column=4, value=p.get('clan_id', '')).border = thin_border
    ws.cell(row=i, column=5, value=p.get('phratry_id', '')).border = thin_border
    ws.cell(row=i, column=6, value=p.get('gender', 'unknown')).border = thin_border
    ws.cell(row=i, column=7, value=','.join(p.get('village_ids', []))).border = thin_border
    ws.cell(row=i, column=8, value=p.get('gudi_village_id', '')).border = thin_border
    ws.cell(row=i, column=9, value='').border = thin_border
    ws.cell(row=i, column=10, value=','.join(p.get('palli', []))).border = thin_border
    ws.cell(row=i, column=11, value=p.get('perma', '')).border = thin_border
    ws.cell(row=i, column=12, value=p.get('karsad', '')).border = thin_border
    ws.cell(row=i, column=13, value=p.get('type', 'main')).border = thin_border
    ws.cell(row=i, column=14, value=p.get('notes', '')).border = thin_border
ws.auto_filter.ref = f'A1:N{len(pens)+1}'
ws.freeze_panes = 'A2'

# Sheet 6: Relationships
ws = wb.create_sheet('Relationships', 5)
write_header(ws, [
    ('type', 14), ('from_pen_id', 25), ('from_pen_name', 30),
    ('to_pen_id', 25), ('to_pen_name', 30), ('details', 60), ('source', 30)
])
for i, r in enumerate(relationships, 2):
    ws.cell(row=i, column=1, value=r['type']).border = thin_border
    ws.cell(row=i, column=2, value=r['from_pen_id']).border = thin_border
    ws.cell(row=i, column=3, value='').border = thin_border
    ws.cell(row=i, column=4, value=r['to_pen_id']).border = thin_border
    ws.cell(row=i, column=5, value='').border = thin_border
    ws.cell(row=i, column=6, value=r.get('details', '')).border = thin_border
    ws.cell(row=i, column=7, value=r.get('source', '')).border = thin_border
ws.auto_filter.ref = f'A1:G{len(relationships)+1}'
ws.freeze_panes = 'A2'

# Save
Path(OUT_XLSX).unlink(missing_ok=True)
wb.save(OUT_XLSX)
print(f'\n✅ Written {OUT_XLSX}')
print(f'  Phratries: {len(phratries)}')
print(f'  Clans: {len(clans)}')
print(f'  Villages: {len(villages)}')
print(f'  Pens: {len(pens)}')
print(f'  Relationships: {len(relationships)}')
