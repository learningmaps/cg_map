"""Build Excel template from clan_gods.json for analyst editing.

Usage:
    python3 scripts/build_excel_template.py
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DATA_PATH = 'data/gods_and_goddesses/clan_gods.json'
OUT_PATH = 'data/gods_and_goddesses/Clan Gods Data.xlsx'

with open(DATA_PATH) as f:
    cg = json.load(f)

wb = Workbook()

header_font = Font(bold=True, color='ffffff', size=11)
header_fill = PatternFill('solid', fgColor='2a6f97')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc'),
)
lock_fill = PatternFill('solid', fgColor='f0f0f0')
readonly_font = Font(color='666666', italic=True)
formula_fill = PatternFill('solid', fgColor='e8f4f8')

PHRATRY_COLORS = {
    'kuhrami_kadiari': 'cc2936',
    'markami': '2a6f97',
    'madvi': '2d6a4f',
    'kawasi': 'e76f00',
}

def style_header(ws, cols):
    for i, name in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

def style_cell(c, readonly=False, formula=False):
    c.alignment = cell_align
    c.border = thin_border
    if readonly:
        c.fill = lock_fill
        c.font = readonly_font
    elif formula:
        c.fill = formula_fill
        c.font = readonly_font

def auto_width(ws, cols, max_width=45):
    for i, name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(name) + 4, 14), max_width)

ws = wb.active
ws.title = 'README'
ws.sheet_properties.tabColor = '2a6f97'

instructions = [
    ['CLAN GODS DATA — Analyst Workbook'],
    [''],
    ['HOW TO USE'],
    ['1. Edit the Villages, Pens, and Relationships sheets below.'],
    ['2. Do NOT change the Phratries or Clans reference sheets without approval.'],
    ['3. When adding new entries, leave ID fields blank — they auto-generate.'],
    ['4. When renaming a pen, add the old name to the aliases column.'],
    ['5. ID columns have dropdowns from reference sheets. Name columns are auto-resolved.'],
    [''],
    ['WORKFLOW'],
    ['  Edit Excel → Run scripts/excel_to_clan_gods.py → clan_gods.json → Map updates'],
    [''],
    ['SHEET REFERENCE'],
    ['  Phratries (lock)  — 4 phratry groups, reference only'],
    ['  Clans (lock)      — 22 clan definitions with phratry mapping'],
    ['  Villages          — Use pen IDs (not names) in main_pen_id and other_pen_ids columns.'],
    ['  Pens              — Use village v_code IDs in gudi_village_id column.'],
    ['  Relationships     — Use pen IDs (not names) in from_pen_id and to_pen_id columns.'],
    [''],
    ['NOTES'],
    ['  • Every pen ID referenced must exist in the Pens sheet.'],
    ['  • Every v_code referenced must exist in the Villages id column.'],
    ['  • Run validation after editing: python3 scripts/excel_to_clan_gods.py --validate'],
    [''],
    ['PHRATRY COLOR KEY'],
    ['  Kuhrami/Kadiari  — RED    (#cc2936)'],
    ['  Markami          — BLUE   (#2a6f97)'],
    ['  Madvi            — GREEN  (#2d6a4f)'],
    ['  Kawasi           — ORANGE (#e76f00)'],
]
for r, row in enumerate(instructions, 1):
    cell = ws.cell(row=r, column=1, value=row[0])
    if r == 1:
        cell.font = Font(bold=True, size=16, color='2a6f97')
    elif row[0].startswith(('PHRATRY', 'SHEET', 'HOW', 'WORKFLOW', 'NOTES')):
        cell.font = Font(bold=True, size=12, color='2a6f97')
ws.column_dimensions['A'].width = 100

# Sheet 2: Phratries (unchanged)
ws = wb.create_sheet('Phratries', 1)
ws.sheet_properties.tabColor = 'cc2936'
phratry_cols = ['id', 'name', 'source', 'clan_ids']
style_header(ws, phratry_cols)
for r, p in enumerate(cg['phratries'], 2):
    for c, key in enumerate(phratry_cols, 1):
        val = p.get(key, '')
        if key == 'clan_ids':
            val = ', '.join(p.get(key, []))
        cell = ws.cell(row=r, column=c, value=val)
        style_cell(cell, readonly=True)
        if key == 'id':
            color = PHRATRY_COLORS.get(val, '888888')
            cell.fill = PatternFill('solid', fgColor=color)
            cell.font = Font(bold=True, color='ffffff')
auto_width(ws, phratry_cols)

# Sheet 3: Clans (unchanged)
ws = wb.create_sheet('Clans', 2)
ws.sheet_properties.tabColor = '3498db'
clan_cols = ['id', 'name', 'phratry_id', 'notes']
style_header(ws, clan_cols)

phratry_dv = DataValidation('list',
    formula1='Phratries!$A$2:$A$%d' % (len(cg['phratries']) + 1),
    allow_blank=True)
phratry_dv.error = 'Must be a valid phratry ID from the Phratries sheet'
ws.add_data_validation(phratry_dv)

for r, clan in enumerate(cg['clans'], 2):
    for c, key in enumerate(clan_cols, 1):
        val = clan.get(key, '')
        cell = ws.cell(row=r, column=c, value=val)
        readonly = key in ('id',)
        style_cell(cell, readonly=readonly)
        if key == 'phratry_id':
            phratry_dv.add(cell)
auto_width(ws, clan_cols)

# Sheet 4: Villages
ws = wb.create_sheet('Villages', 3)
ws.sheet_properties.tabColor = '27ae60'
village_cols = ['name', 'aliases', 'phratry_id', 'clan_id', 'main_pen_id',
                'main_pen_name', 'other_pen_ids', 'notes', 'population_info', 'id']
style_header(ws, village_cols)

phratry_dv2 = DataValidation('list',
    formula1='Phratries!$A$2:$A$%d' % (len(cg['phratries']) + 1), allow_blank=True)
ws.add_data_validation(phratry_dv2)

clan_dv = DataValidation('list',
    formula1='Clans!$A$2:$A$%d' % (len(cg['clans']) + 1), allow_blank=True)
ws.add_data_validation(clan_dv)

pen_id_dv = DataValidation('list',
    formula1='Pens!$A$2:$A$%d' % (len(cg['pens']) + 1), allow_blank=True)
pen_id_dv.error = 'Must be a valid pen ID from the Pens sheet'
ws.add_data_validation(pen_id_dv)

for r, v in enumerate(cg['villages'], 2):
    main_pen_id = v.get('main_pen_id') or ''
    sub_ids = []
    for sid in v.get('subordinate_pen_ids', []):
        sub_ids.append(sid)
    vals = {
        'name': v['name'],
        'aliases': ', '.join(v.get('aliases', [])),
        'phratry_id': v['phratry_id'],
        'clan_id': v['clan_id'],
        'main_pen_id': main_pen_id,
        'main_pen_name': '',
        'other_pen_ids': ', '.join(sub_ids),
        'notes': v.get('notes', ''),
        'population_info': v.get('population_info', ''),
        'id': v.get('id') or '',
    }
    for c, key in enumerate(village_cols, 1):
        cell = ws.cell(row=r, column=c, value=vals[key])
        readonly = key in ('id', 'main_pen_name')
        formula = key == 'main_pen_name'
        style_cell(cell, readonly=readonly, formula=formula)
        if formula:
            cell.value = f'=IF(E{r}="","",XLOOKUP(E{r}, Pens!$A:$A, Pens!$B:$B))'
        if key == 'phratry_id':
            phratry_dv2.add(cell)
        elif key == 'clan_id':
            clan_dv.add(cell)
        elif key == 'main_pen_id':
            pen_id_dv.add(cell)
auto_width(ws, village_cols)

# Sheet 5: Pens
ws = wb.create_sheet('Pens', 4)
ws.sheet_properties.tabColor = 'e67e22'
pen_cols = ['id', 'name', 'aliases', 'clan_id', 'phratry_id', 'gender',
            'gudi_village_id', 'gudi_village_name', 'palli', 'perma', 'karsad', 'type', 'notes']
style_header(ws, pen_cols)

phratry_dv3 = DataValidation('list',
    formula1='Phratries!$A$2:$A$%d' % (len(cg['phratries']) + 1), allow_blank=True)
ws.add_data_validation(phratry_dv3)

clan_dv2 = DataValidation('list',
    formula1='Clans!$A$2:$A$%d' % (len(cg['clans']) + 1), allow_blank=True)
ws.add_data_validation(clan_dv2)

gender_dv = DataValidation('list', formula1='"male,female,unknown"', allow_blank=True)
gender_dv.error = 'Must be male, female, or unknown'
ws.add_data_validation(gender_dv)

type_dv = DataValidation('list', formula1='"main,subordinate"', allow_blank=True)
ws.add_data_validation(type_dv)

# Build list of village v_codes for dropdown
village_ids = sorted(set(v.get('id') for v in cg['villages'] if v.get('id')))
village_id_range = 'Villages!$J$2:$J$%d' % (len(cg['villages']) + 1)
village_id_dv = DataValidation('list',
    formula1=village_id_range, allow_blank=True)
village_id_dv.error = 'Must be a valid village v_code from Villages id column'
ws.add_data_validation(village_id_dv)

for r, pen in enumerate(cg['pens'], 2):
    vals = {
        'id': pen['id'],
        'name': pen['name'],
        'aliases': ', '.join(pen.get('aliases', [])),
        'clan_id': pen.get('clan_id') or '',
        'phratry_id': pen.get('phratry_id') or '',
        'gender': pen.get('gender', 'unknown'),
        'gudi_village_id': pen.get('gudi_village_id') or '',
        'gudi_village_name': '',
        'palli': ', '.join(pen.get('palli', [])),
        'perma': pen.get('perma', ''),
        'karsad': pen.get('karsad', ''),
        'type': pen.get('type', 'main'),
        'notes': pen.get('notes', ''),
    }
    for c, key in enumerate(pen_cols, 1):
        cell = ws.cell(row=r, column=c, value=vals[key])
        readonly = key in ('id', 'gudi_village_name')
        formula = key == 'gudi_village_name'
        style_cell(cell, readonly=readonly, formula=formula)
        if formula:
            cell.value = f'=IF(G{r}="","",XLOOKUP(G{r}, Villages!$J:$J, Villages!$A:$A))'
        if key == 'phratry_id':
            phratry_dv3.add(cell)
        elif key == 'clan_id':
            clan_dv2.add(cell)
        elif key == 'gender':
            gender_dv.add(cell)
        elif key == 'type':
            type_dv.add(cell)
        elif key == 'gudi_village_id':
            village_id_dv.add(cell)
auto_width(ws, pen_cols)

# Sheet 6: Relationships
ws = wb.create_sheet('Relationships', 5)
ws.sheet_properties.tabColor = '9b59b6'
rel_cols = ['type', 'from_pen_id', 'from_pen_name', 'to_pen_id', 'to_pen_name', 'details', 'source']
style_header(ws, rel_cols)

rel_type_dv = DataValidation('list',
    formula1='"spouse,sibling,parent,child,ghar_jamai"', allow_blank=True)
ws.add_data_validation(rel_type_dv)

pen_id_dv2 = DataValidation('list',
    formula1='Pens!$A$2:$A$%d' % (len(cg['pens']) + 1), allow_blank=True)
ws.add_data_validation(pen_id_dv2)

for r, rel in enumerate(cg['relationships'], 2):
    vals = {
        'type': rel['type'],
        'from_pen_id': rel['from_pen_id'],
        'from_pen_name': '',
        'to_pen_id': rel['to_pen_id'],
        'to_pen_name': '',
        'details': rel.get('details', ''),
        'source': rel.get('source', ''),
    }
    for c, key in enumerate(rel_cols, 1):
        cell = ws.cell(row=r, column=c, value=vals[key])
        formula = key in ('from_pen_name', 'to_pen_name')
        style_cell(cell, formula=formula)
        if key == 'from_pen_name':
            cell.value = f'=IF(B{r}="","",XLOOKUP(B{r}, Pens!$A:$A, Pens!$B:$B))'
        elif key == 'to_pen_name':
            cell.value = f'=IF(D{r}="","",XLOOKUP(D{r}, Pens!$A:$A, Pens!$B:$B))'
        elif key == 'type':
            rel_type_dv.add(cell)
        elif key in ('from_pen_id', 'to_pen_id'):
            pen_id_dv2.add(cell)
auto_width(ws, rel_cols)

for name in ['Villages', 'Pens', 'Relationships', 'Clans']:
    wb[name].freeze_panes = 'A2'

wb.save(OUT_PATH)
print(f"✓ Wrote {OUT_PATH}")
print(f"  Phratries: {len(cg['phratries'])}")
print(f"  Clans: {len(cg['clans'])}")
print(f"  Villages: {len(cg['villages'])}")
print(f"  Pens: {len(cg['pens'])}")
print(f"  Relationships: {len(cg['relationships'])}")
