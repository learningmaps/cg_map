import docx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

doc = docx.Document('data/Extra Data/gods_and_goddesses/List of Clan Gods and Villages.docx')
tables = doc.tables

wb = Workbook()
header_font = Font(bold=True, color='ffffff', size=11)
header_fill = PatternFill('solid', fgColor='2a6f97')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc'),
)

PEN_COLS = [
    'name', 'aliases', 'gender', 'clan', 'phratry',
    'village', 'village_details', 'palli', 'perma', 'karsad',
    'type', 'spouse', 'siblings', 'children', 'parent', 'notes'
]

ws = wb.active
ws.title = 'Pens'
for i, col_name in enumerate(PEN_COLS, 1):
    c = ws.cell(row=1, column=i, value=col_name)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = thin_border

row_num = 2

def add_pen(ws, row_num, name, gender, clan, phratry, village,
            spouse=None, siblings=None, children=None, parent=None,
            notes=None, type='main', aliases=None, palli=None):
    vals = [
        name,
        ', '.join(aliases) if aliases else '',
        gender,
        clan,
        phratry,
        village,
        '',
        ', '.join(palli) if palli else '',
        '',
        '',
        type,
        ', '.join(spouse) if spouse else '',
        ', '.join(siblings) if siblings else '',
        ', '.join(children) if children else '',
        ', '.join(parent) if parent else '',
        notes or '',
    ]
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.alignment = cell_align
        cell.border = thin_border
        if c == 1:
            cell.font = Font(bold=True)
    return row_num + 1

def sp(s): return s.strip()

# ─────────────────────────────────────────────
# Phratry 1: Kuhrami / Kadiari  (Table 0)
# Clan: Kunjam
# ─────────────────────────────────────────────
t0 = [[sp(cell.text) for cell in row.cells] for row in tables[0].rows]

# Manual extraction per DOCX table 0
pens = [
    # (name, gender, clan, phratry, village, spouse, siblings, children, parent, notes, type, aliases, palli)
    ('Vedmo Moitor', 'male', 'Kunjam', 'Kuhrami/Kadiari', 'Daler (near Marh)',
     None, ['Urru Moitor'], None, None, 'Younger brother of Urru Moitor', 'main', None, None),

    ('Urru Moitor', 'male', 'Kunjam', 'Kuhrami/Kadiari', 'Omalwar/Samalwar',
     ['Urru Dokri'], ['Perambhoi Raj'], None, None,
     'Gudi in Omalwar. Urru Moitor and Urru Dokri have their gudi in Omalwar', 'main', None, None),

    ('Perambhoi Raj', 'male', 'Kunjam', 'Kuhrami/Kadiari', 'Omalwar/Samalwar',
     None, ['Urru Moitor'], None, None, 'Younger brother of Urru Moitor', 'subordinate', None, None),

    ('Punga Rungi', 'female', 'Kunjam', 'Kuhrami/Kadiari', 'Benpal/Bayampal',
     ['Mai Sunga'], None, None, None,
     'Telam wife, married to Mai Sunga Kunjam pen husband, now ghar jamai in Bhansi', 'main', None, None),

    ('Chichur Unga', 'male', 'Kunjam', 'Kuhrami/Kadiari', 'Hiroli',
     None, None, None, None,
     'Bachelor. To get married to Pal Urre, d/o Andal Kosa Madvi pen of Cholnar', 'main', ['Chichur Urra'], None),

    ('Gariaram', 'male', 'Kunjam', 'Kuhrami/Kadiari', 'Tamodi',
     None, ['Chichur Unga'], None, None,
     'Younger brother of Chichur Unga', 'main', None, None),

    ('Hadma Raj', 'male', 'Kunjam', 'Kuhrami/Kadiari', 'Palnar',
     None, None, None, ['Urru Moitor'],
     's/o Urru Moitor', 'main', None, None),

    ('Hadma Raj', 'male', 'Kunjam', 'Kuhrami/Kadiari', 'Kuwe',
     None, None, None, ['Urru Moitor'],
     's/o Urru Moitor (same pen as Palnar)', 'main', None, None),

    ('Kundel Ganga', 'male', 'Kunjam', 'Kuhrami/Kadiari', 'Shyamgiri',
     None, ['Hadma Raj'], None, ['Urru Moitor'],
     'Also called Jalsingo. Brother of Hadma Raj, s/o Urru Moitor', 'main', ['Jalsingo'], None),

    # Gongpal - unnamed son, use descriptive name
    ('(unnamed son of Punga Rungi and Mai Sunga)', 'male', 'Kunjam', 'Kuhrami/Kadiari', 'Gongpal',
     None, None, None, ['Punga Rungi', 'Mai Sunga'],
     'Name not given in source', 'main', None, None),

    # Kunjampara - unnamed son
    ('(unnamed son of Punga Rungi and Mai Sunga)', 'male', 'Kunjam', 'Kuhrami/Kadiari', 'Kunjampara, near Ganjenar',
     None, None, None, ['Punga Rungi', 'Mai Sunga'],
     'Name not given in source', 'main', None, None),

    # Kodoli - uncertain parentage
    ('(unnamed pen)', 'male', 'Kunjam', 'Kuhrami/Kadiari', 'Kodoli',
     None, None, None, None,
     'Either a son of Punga Rungi & Mai Sunga or a son of Urru Moitor', 'main', None, None),
]

for p in pens:
    row_num = add_pen(ws, row_num, *p)
print(f"Kuhrami/Kadiari: {len(pens)} pens")

# ─────────────────────────────────────────────
# Phratry 2: Markami Kutumb  (Table 1)
# Many clans
# ─────────────────────────────────────────────
t1 = [[sp(cell.text) for cell in row.cells] for row in tables[1].rows]

pens = [
    ('Iram Raj', 'male', 'Markami', 'Markami Kutumb', 'Markagudem',
     None, None, None, None, 'Also called Irma Raj', 'main', ['Irma Raj'], None),

    ('Hunga Moitor', 'male', 'Kalmu/Karma', 'Markami Kutumb', 'Dugeli',
     ['Gujje Dokri', 'Murde Moyo', 'Ir Sunge'], None, None, None,
     'At the Cholnar karsad there was also Budi Moyo who had come with Hunga Moitor', 'main', None, None),

    ('Gujje Dokri', 'female', 'Kalmu/Karma', 'Markami Kutumb', 'Pharaspal',
     ['Hunga Moitor'], None, None, None,
     'Made of irirom mada in Songunda above Kondanpal. Married to Hunga Moitor in Dugeli. Visit each other at karsads.', 'main', None, None),

    ('Sungal', 'male', 'Karma', 'Markami Kutumb', 'Dhurli',
     None, None, None, None, '', 'main', None, None),

    ('Bhimaraj', 'male', 'Bhogam/Chote Telam', 'Markami Kutumb', 'Madpal',
     ['Urru Ponde', 'Dol Mutte'], ['Biriya Bhima', 'Chaikud Bhima', 'Huru Mara (needs clarification)'], None, None,
     'Brothers of Huru Mara of Bhansi (needs clarification)', 'main', None, None),

    ('Biriya Bhima', 'male', 'Chote Telam/Bhogam', 'Markami Kutumb', 'Madpal',
     ['Mudde Dokri'], ['Bhimaraj', 'Chaikud Bhima'], None, None,
     'Mudde Dokri is d/o Muddaraj. Biriya Bhima, Mudde Dokri and Urru Ponde counted as Chote Telam. Bhogams also known as Chote Telams.', 'main', None, None),

    ('Bomul Ungal', 'male', 'Bhogam', 'Markami Kutumb', 'Mirtulnar/Midkulnar',
     None, None, None, ['Biriya Bhima'],
     's/o Biriya Bhima', 'main', None, None),

    ('Godar Bhima', 'male', 'Chote Telam/Bhogam', 'Markami Kutumb', 'Manganar',
     ['Bhime (d/o Gadi Kama Oyam)'], ['Chaikut Bhima'], None, None,
     'Chaikut = small, Godar = one who crossed Godavari. Son across Godavari; daughter Inge Dokri has pen in Kuper. Palli: Palnar, Teknar, Chitalur, Kawalnar, Mangnar, Midkulnar, bits of Pharaspal, Kesapur, Turparas, Idodpara. Offerings at Diwad and Bijja Pandum.', 'main', None,
     ['Palnar', 'Teknar', 'Chitalur', 'Kawalnar', 'Mangnar', 'Midkulnar', 'Pharaspal', 'Kesapur', 'Turparas', 'Idodpara']),

    ('Urru Mara (Huru Mara)', 'male', 'Bade Telam', 'Markami Kutumb', 'Bhansi',
     ['Urru Dokri', 'Mawe Lungo'], None, ['Lug Unga', 'Mai Sunga (ghar jamai)'], None,
     'Mawe Lungo is d/o Nandraj and Mawli. Also cross-referenced in Table 0 (Kunjam phratry). Gudi in Omalwar with Urru Dokri. Palli: Jhirka, Dokometta, Dhanora, Tumnar, Bhansi.', 'main', ['Hura Mara', 'Huru Mara'],
     ['Jhirka', 'Dokometta', 'Dhanora', 'Tumnar', 'Bhansi']),

    ('Lug Unga', 'male', 'Bade Telam', 'Markami Kutumb', 'Bhansi',
     None, None, None, ['Urru Mara', 'Mawe Lungo'],
     's/o Huru Mara and Mawe Lungo', 'subordinate', None, None),

    ('Mai Sunga', 'male', 'Bade Telam', 'Markami Kutumb', 'Bhansi',
     ['Punga Rungi'], None, None, None,
     'Ghar jamai, married to Punga Rungi d/o Huru Mara and Urru Dokri. Has his palli and dhaniya in Jaraloha.', 'subordinate', None, None),

    ('Godar Hunga', 'male', 'Bade Telam', 'Markami Kutumb', 'Jhirka',
     None, None, None, ['Huru Mara', 'Katte Bodke'],
     's/o Huru Mara and Katte Bodke. See case study of Jhirka on phallis.', 'main', None, None),

    ('Daro Moitor (Darrem Modka)', 'male', 'Telam', 'Markami Kutumb', 'Renganar',
     None, ['Huru Mara'], None, None,
     'Elder brother of Huru Mara of Bhansi. Palli: Chandenar, Renganar.', 'main', ['Darrem Modka'],
     ['Chandenar', 'Renganar']),

    ('Raibandal', 'male', 'Kadiyam', 'Markami Kutumb', 'Kondapal',
     ['Akaluru Dokri'], None, None, None, '', 'main', None, None),

    ('Kohla Kosu', 'male', 'Kadiyam', 'Markami Kutumb', 'Metapal',
     None, None, None, ['Raibandal'],
     's/o Raibandal', 'main', None, None),

    ('Peda Hadma', 'male', 'Midiyam', 'Markami Kutumb', 'Gumiyapal (Guyempad)',
     None, ['Chinna Hadma'], None, None, '', 'main', None, None),

    ('Chinna Hadma', 'male', 'Midiyam', 'Markami Kutumb', 'Gumiyapal (Guyempad)',
     None, ['Peda Hadma'], None, None, '', 'main', None, None),

    ('Murde Muyal', 'male', 'Undam', 'Markami Kutumb', 'Vengpal (Vengur)',
     None, ['Hunga Bhimal'], None, None, '', 'main', None, None),

    ('Hunga Bhimal', 'male', 'Undam', 'Markami Kutumb', 'Vengpal (Vengur)',
     None, ['Murde Muyal'], None, None, '', 'main', None, None),

    ('Pind Hadmal', 'male', 'Tati', 'Markami Kutumb', 'Tikanpal',
     None, None, None, None, '', 'main', None, None),

    ('Bhum Iriyal', 'male', 'Kadti', 'Markami Kutumb', 'Madadi',
     None, None, None, None,
     'His daughter Kohli Dokri is married to Gaddi Kama, Oyam Pen of Pandewar', 'main', None, None),

    ('Pal Hadmal', 'male', 'Rengo', 'Markami Kutumb', 'Jhadka (Orcha Block)',
     None, None, None, None, '', 'main', None, None),
]

# Note: Table 1 also mentions some villages without specific named pens:
# Baregunda (Row 13) - Telam village, no specific pen mentioned
# Pendvela, Alnar (Row 15) - Icham, no pen name
# Katural, Pumbad (Row 16) - Punem, no pen name
# Skipping these as they have no named pen

for p in pens:
    row_num = add_pen(ws, row_num, *p)
print(f"Markami Kutumb: {len(pens)} pens")

# ─────────────────────────────────────────────
# Phratry 3: Madvi  (Tables 2 + 3)
# ─────────────────────────────────────────────
t2 = [[sp(cell.text) for cell in row.cells] for row in tables[2].rows]

pens = [
    ('Muddaraj', 'male', 'Oyam', 'Madvi', 'Vechapal',
     None, None, ['Mudde Dokri'], None,
     'Father of Mudde Dokri who is married to Biriya Bhima (Chote Telam) of Madpal. Phalli extends to Madpal.', 'main', None, None),

    ('Vidi Iriyal Panda Hadma', 'male', 'Oyam', 'Madvi', 'Kesapur',
     None, ['Gaddi Kama'], None, None,
     'Third and youngest unmarried brother of Gaddi Kama', 'main', None, None),

    ('Gaddi Kama', 'male', 'Oyam', 'Madvi', 'Pandewar',
     ['Murke Nango', 'Kohli Dokri'], ['Vidi Iriyal Panda Hadma', 'Bade Vidi Iriyal', 'Hadma Iriyal'], ['Gal Dullo', 'Gal Bomda', 'Hingal Devo', 'Chenna Kama', 'Ghanta Kama', 'Ukud Kama'], None,
     'Also written as Ghadi Kama. Children with Murke Nango: Gal Dullo (Etlapad), Gal Bomda (Etlapad), Hingal Devo (Korre Kongdam), 1 more boy. With Kohli Dokri: Chenna Kama, Ghanta Kama, Ukud Kama (all in Pandewar). Murke Nango is d/o Huru Mara of Bhansi. Kohli Dokri is d/o Bhum Iriya (Kadti pen, Madadi).', 'main', ['Ghadi Kama'], None),

    ('Gal Dullo', 'male', 'Oyam', 'Madvi', 'Etlapad',
     None, ['Gal Bomda'], None, ['Gaddi Kama'],
     'Child of Gaddi Kama', 'main', None, None),

    ('Gal Bomda', 'male', 'Oyam', 'Madvi', 'Etlapad',
     None, ['Gal Dullo'], None, ['Gaddi Kama'],
     'Child of Gaddi Kama', 'main', None, None),

    ('Bade Vidi Iriyal', 'male', 'Oyam', 'Madvi', 'Rekavaya',
     None, ['Gaddi Kama', 'Hadma Iriyal'], None, None,
     'Second brother of Gaddi Kama. No sisters.', 'main', None, None),

    ('Hadma Iriyal', 'male', 'Oyam', 'Madvi', 'Kamkajojor',
     None, ['Gaddi Kama', 'Bade Vidi Iriyal'], None, None,
     'Brother of Gaddi Kama', 'main', None, None),

    ('Bande Boyo', 'male', 'Oyam', 'Madvi', 'Pidiya',
     None, ['Gaddi Kama'], None, None,
     'May or may not be brother of Gaddi Kama', 'main', None, None),
]

for p in pens:
    row_num = add_pen(ws, row_num, *p)
print(f"Madvi (Oyam): {len(pens)} pens")

# Table 3: Madvi contd (other clans)
t3 = [[sp(cell.text) for cell in row.cells] for row in tables[3].rows]

pens = [
    ('Vange (Vunge) Dokri', 'female', 'Hemla', 'Madvi', 'Pinkonda',
     ['Nanga Bhima'], None, None, None,
     'Married to Nanga Bhima in Gangalur village', 'main', ['Vunge Dokri'], None),

    ('Uraal Gundal', 'male', 'Tamo', 'Madvi', 'Tamirguda',
     None, None, None, None, '', 'main', None, None),

    ('Barra Bujja', 'male', 'Padami', 'Madvi', 'Vengur',
     None, None, None, None, '', 'main', None, None),

    ('Morka Moitor', 'male', 'Ujji/Dodi', 'Madvi', 'Tadopadar',
     None, None, None, None, '', 'main', None, None),

    ('Markaraj (Marka Moitor)', 'male', 'Barse', 'Madvi', 'Kamalnar/Kamalur',
     ['Hinge Dokri'], None, None, None,
     'Palli: Dantewada jail, Markanar, Kumharras, Matenar, Kondoli (son lives), Gamawada (son also called Markaraj), Ganjenar (son/bro), Mulasnar (Gumorongo s/o Markaraj), Balod, Kuper (chud inga, bro), Veesro, Kundundel', 'main', ['Marka Moitor'], None),

    ('Handa (Andal) Kosa', 'male', 'Madvi', 'Madvi', 'Cholnar',
     ['Pal Hadme (d/o Hirma Raj of Markamiras)'], ['Chudanda'], ['Palenda'], None,
     'Married to Pal Hadme, d/o Hirma Raj of Markamiras', 'main', ['Andal Kosa'], None),

    ('Palenda', 'male', 'Madvi', 'Madvi', 'Bododi',
     ['Pole Koso'], None, None, ['Handa Kosa'],
     's/o Handa Kosa', 'main', None, None),

    ('Chudanda', 'male', 'Madvi', 'Madvi', 'Pondum',
     None, ['Handa Kosa'], None, None,
     'Younger brother of Handa Kosa', 'main', None, None),
]

for p in pens:
    row_num = add_pen(ws, row_num, *p)
print(f"Madvi (other): {len(pens)} pens")

# ─────────────────────────────────────────────
# Phratry 4: Kawasi  (Table 4)
# ─────────────────────────────────────────────
t4 = [[sp(cell.text) for cell in row.cells] for row in tables[4].rows]

pens = [
    ('Babo Harma Pen', 'male', 'Kawasi', 'Kawasi', 'Jawanga',
     None, None, None, None,
     'Source: Grigson', 'main', None, None),
]

for p in pens:
    row_num = add_pen(ws, row_num, *p)
print(f"Kawasi: {len(pens)} pen")

# ─────────────────────────────────────────────
# Phratry 5: Sodi  (Table 5)
# ─────────────────────────────────────────────
t5 = [[sp(cell.text) for cell in row.cells] for row in tables[5].rows]

pens = [
    ('Tal Muttai', 'male', 'Sodi', 'Sodi', 'Kamkanar',
     None, None, None, None,
     '', 'main', None, None),
]

for p in pens:
    row_num = add_pen(ws, row_num, *p)
print(f"Sodi: {len(pens)} pen")

# ─────────────────────────────────────────────
# Formatting & column widths
# ─────────────────────────────────────────────
col_widths = {1: 30, 2: 25, 3: 10, 4: 22, 5: 22, 6: 28, 7: 35, 8: 45, 9: 18, 10: 18, 11: 14, 12: 30, 13: 30, 14: 28, 15: 28, 16: 60}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:P{row_num - 1}'

# ─────────────────────────────────────────────
# Data Dictionary sheet
# ─────────────────────────────────────────────
ws2 = wb.create_sheet('Data Dictionary', 1)
dict_cols = ['Field', 'Description', 'Example', 'Required', 'Rules']
dict_header_fill = PatternFill('solid', fgColor='3498db')
for i, name in enumerate(dict_cols, 1):
    c = ws2.cell(row=1, column=i, value=name)
    c.font = Font(bold=True, color='ffffff', size=11)
    c.fill = dict_header_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

dict_data = [
    ['name', 'Name of the pen (deity/god). Parentheses indicate name not known.', 'Godar Hunga', 'Yes', 'Will be used to auto-generate a pen ID. Descriptive names used when actual name unknown.'],
    ['aliases', 'Alternate names or spellings found in source.', 'Godar Hunga, Godar Hunga', 'No', 'Comma-separated.'],
    ['gender', 'Gender of the pen.', 'male', 'Yes', 'male / female / unknown. Inferred from context (kinship terms, dokri suffix, etc.).'],
    ['clan', 'Clan this pen belongs to.', 'Bade Telam', 'Yes if known', 'From the source DOCX Clan column. Some pens have dual clan names (e.g., Bhogam/Chote Telam).'],
    ['phratry', 'Phratry grouping.', 'Markami Kutumb', 'Yes if known', 'Kuhrami/Kadiari, Markami Kutumb, Madvi, Kawasi, Sodi. Matches section headers.'],
    ['village', 'Gudi village where the pen shrine is located.', 'Jhirka', 'No', 'Free text from the source DOCX.'],
    ['village_details', 'Extra info about the gudi village.', 'Near Jhirka waterfall', 'No', 'Free text.'],
    ['palli', 'Territory of influence - villages where this pen is worshipped.', 'Dokometta, Dhanora', 'No', 'Comma-separated. Extracted from Remarks column when explicitly listed.'],
    ['perma', 'Name of the priest (perma).', '', 'No', 'Not available in source DOCX for any entry.'],
    ['karsad', 'Annual festival day.', '', 'No', 'Not available in source DOCX for any entry.'],
    ['type', 'Main or subordinate pen.', 'main', 'Yes', 'main / subordinate. Pens explicitly listed as children, ghar jamai, or unnamed relatives marked as subordinate.'],
    ['spouse', 'Spouse(s). Prefix ghar_jamai: for ghar jamai.', 'Punga Rungi', 'No', 'Pen name(s) from Relations column. ghar_jamai noted in parentheses.'],
    ['siblings', 'Sibling pens.', 'Huru Mara, Daro Moitor', 'No', 'Comma-separated. Extracted from "brother of" statements.'],
    ['children', 'Children of this pen.', 'Lug Unga, Punga Rungi', 'No', 'Comma-separated. From s/o or children listed in source.'],
    ['parent', 'Parent of this pen.', 'Huru Mara', 'No', 'From s/o statements. Multiple parents separated by comma.'],
    ['notes', 'Any additional notes from the source DOCX.', 's/o Huru Mara and Katta Bodke; Bachelor', 'No', 'Free text. Includes info from Remarks column and additional context.'],
]
for r, row_data in enumerate(dict_data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border
        if c == 1:
            cell.font = Font(bold=True)
ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 60
ws2.column_dimensions['C'].width = 45
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 60
ws2.freeze_panes = 'A2'

# Save
out_path = 'data/gods_and_goddesses/Clan Gods Data - Simple.xlsx'
Path(out_path).unlink(missing_ok=True)
wb.save(out_path)
print(f'\n✅ Saved {out_path}')
print(f'Total rows in Pens sheet: {row_num - 2} (incl. section headers)')
print(f'Data Dictionary: {len(dict_data)} fields')
