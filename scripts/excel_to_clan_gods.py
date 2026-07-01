"""Convert analyst-edited Excel workbook back to clan_gods.json.

Usage:
    python3 scripts/excel_to_clan_gods.py              # Convert Excel -> JSON
    python3 scripts/excel_to_clan_gods.py --validate    # Validate only, no output
"""
import json
import re
import sys
from openpyxl import load_workbook

EXCEL_PATH = 'data/gods_and_goddesses/Clan Gods Data.xlsx'
OUT_PATH = 'data/gods_and_goddesses/clan_gods.json'


def make_id(name):
    if not name:
        return None
    name = name.strip()
    name = name.replace('\n', ' ').replace('\r', ' ')
    n = name.lower()
    n = re.sub(r'[^a-z0-9_\s]', '', n)
    n = re.sub(r'\s+', '_', n.strip())
    return n


def parse_aliases(val):
    if not val:
        return []
    return [a.strip() for a in val.split(',') if a.strip()]


def parse_comma_list(val):
    if not val:
        return []
    return [v.strip() for v in val.split(',') if v.strip()]


def infer_gender(name):
    if not name:
        return 'unknown'
    female_indicators = ['dokri', 'dokrar', 'moyo', 'lungo', 'rungi', 'ponde', 'nango', 'bhime', 'bodke', 'mutte']
    for fi in female_indicators:
        if fi in name.lower():
            return 'female'
    return 'male'


def load_excel():
    wb = load_workbook(EXCEL_PATH)
    data = {}

    # Phratries
    ws = wb['Phratries']
    phratries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip():
            phratries.append({
                'id': str(row[0]).strip(),
                'name': str(row[1] or '').strip(),
                'source': str(row[2] or '').strip(),
                'clan_ids': [c.strip() for c in str(row[3] or '').split(',') if c.strip()],
            })
    data['phratries'] = phratries

    # Clans
    ws = wb['Clans']
    clans = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip():
            clans.append({
                'id': str(row[0]).strip(),
                'name': str(row[1] or '').strip(),
                'phratry_id': str(row[2] or '').strip(),
                'notes': str(row[3] or '').strip(),
            })
    data['clans'] = clans

    # Villages
    # Columns: name, aliases, phratry_id, clan_id, main_pen_id, main_pen_name(formula), other_pen_ids, notes, population_info, id
    ws = wb['Villages']
    villages = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not str(row[0]).strip():
            continue
        main_pen_id = str(row[4] or '').strip() if row[4] else ''
        v = {
            'name': str(row[0]).strip(),
            'aliases': parse_aliases(str(row[1] or '')),
            'phratry_id': str(row[2] or '').strip(),
            'clan_id': str(row[3] or '').strip(),
            'main_pen_id': main_pen_id if main_pen_id else '',
            'subordinate_pen_ids': parse_comma_list(str(row[6] or '')),
            'notes': str(row[7] or '').strip(),
            'population_info': str(row[8] or '').strip(),
            'id': str(row[9] or '').strip() if row[9] else None,
        }
        if v['id'] == '':
            v['id'] = None
        villages.append(v)
    data['villages'] = villages

    # Pens
    # Columns: id, name, aliases, clan_id, phratry_id, gender, village_ids, gudi_village_id, gudi_village_name(formula), palli, perma, karsad, type, notes
    ws = wb['Pens']
    pens = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1] or not str(row[1]).strip():
            continue
        pen_id = str(row[0] or '').strip() if row[0] else ''
        if not pen_id:
            pen_id = make_id(str(row[1]).strip())
        vids = parse_comma_list(str(row[6] or ''))
        gvi = str(row[7] or '').strip() if row[7] else ''
        pens.append({
            'id': pen_id,
            'name': str(row[1]).strip(),
            'aliases': parse_aliases(str(row[2] or '')),
            'clan_id': str(row[3] or '').strip() or None,
            'phratry_id': str(row[4] or '').strip() or None,
            'gender': str(row[5] or infer_gender(str(row[1] or ''))).strip(),
            'village_ids': vids,
            'gudi_village_id': gvi or None,
            'palli': parse_comma_list(str(row[9] or '')),
            'perma': str(row[10] or '').strip(),
            'karsad': str(row[11] or '').strip(),
            'type': str(row[12] or 'main').strip(),
            'notes': str(row[13] or '').strip(),
        })
    data['pens'] = pens

    # Relationships
    # Columns: type, from_pen_id, from_pen_name(formula), to_pen_id, to_pen_name(formula), details, source
    ws = wb['Relationships']
    relationships = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not str(row[0]).strip():
            continue
        from_id = str(row[1] or '').strip()
        to_id = str(row[3] or '').strip()
        if not from_id or not to_id:
            continue
        relationships.append({
            'type': str(row[0]).strip(),
            'from_pen_id': from_id,
            'to_pen_id': to_id,
            'details': str(row[5] or '').strip(),
            'source': str(row[6] or '').strip(),
        })
    data['relationships'] = relationships

    wb.close()
    return data


def resolve_gudi_village_names(data, existing=None):
    """Look up village names from v_code IDs and set gudi_village / villages fields."""
    vcode_to_name = {}
    for v in data['villages']:
        vid = v.get('id')
        if vid:
            vcode_to_name[vid] = v['name']

    old_pen_map = {}
    if existing:
        for p in existing.get('pens', []):
            old_pen_map[p['id']] = p

    for pen in data['pens']:
        # Resolve all village IDs to names
        village_names = []
        for vid in pen.get('village_ids', []):
            if vid in vcode_to_name:
                village_names.append(vcode_to_name[vid])
        pen['villages'] = village_names

        # Primary village (backward compat)
        gvi = pen.get('gudi_village_id')
        if gvi and gvi in vcode_to_name:
            pen['gudi_village'] = vcode_to_name[gvi]
        elif gvi:
            pen['gudi_village'] = f'UNKNOWN_VILLAGE_{gvi}'
        else:
            old_pen = old_pen_map.get(pen['id'])
            if old_pen and old_pen.get('gudi_village'):
                pen['gudi_village'] = old_pen['gudi_village']
            else:
                pen['gudi_village'] = None

    return data


def find_unreferenced(data):
    issues = []
    pen_ids = {p['id'] for p in data['pens']}
    clan_ids = {c['id'] for c in data['clans']}
    phratry_ids = {p['id'] for p in data['phratries']}
    village_ids = {v['id'] for v in data['villages'] if v.get('id')}

    for i, v in enumerate(data['villages']):
        if v['phratry_id'] and v['phratry_id'] not in phratry_ids:
            issues.append(f'Village[{i}] "{v["name"]}": unknown phratry "{v["phratry_id"]}"')
        if v['clan_id'] and v['clan_id'] not in clan_ids:
            issues.append(f'Village[{i}] "{v["name"]}": unknown clan "{v["clan_id"]}"')
        if v['main_pen_id'] and v['main_pen_id'] not in pen_ids:
            issues.append(f'Village[{i}] "{v["name"]}": unknown main_pen_id "{v["main_pen_id"]}"')
        for sub_id in v.get('subordinate_pen_ids', []):
            if sub_id and sub_id not in pen_ids:
                issues.append(f'Village[{i}] "{v["name"]}": unknown subordinate pen ID "{sub_id}"')

    for i, p in enumerate(data['pens']):
        gvi = p.get('gudi_village_id')
        if gvi and gvi not in village_ids:
            issues.append(f'Pen[{i}] "{p["name"]}": unknown gudi_village_id "{gvi}"')
        for vid in p.get('village_ids', []):
            if vid and vid not in village_ids:
                issues.append(f'Pen[{i}] "{p["name"]}": unknown village_id "{vid}"')

    for i, r in enumerate(data['relationships']):
        if r['from_pen_id'] not in pen_ids:
            issues.append(f'Relation[{i}]: unknown from_pen_id "{r["from_pen_id"]}"')
        if r['to_pen_id'] not in pen_ids:
            issues.append(f'Relation[{i}]: unknown to_pen_id "{r["to_pen_id"]}"')

    return issues


def build_phratries_list(data):
    return data['phratries']


def build_metadata():
    return {
        'title': 'List of Clan Gods and Villages in Dantewada',
        'source_file': 'Clan Gods Data.xlsx (analyst workbook)',
        'extraction_date': '2026-06-29',
        'phratry_colors': {
            'kuhrami_kadiari': {'fill': '#cc2936', 'label': 'Kuhrami/Kadiari'},
            'markami': {'fill': '#2a6f97', 'label': 'Markami Kutumb'},
            'madvi': {'fill': '#2d6a4f', 'label': 'Madvi'},
            'kawasi': {'fill': '#e76f00', 'label': 'Kawasi'},
        },
    }


def main():
    validate_only = '--validate' in sys.argv

    data = load_excel()

    try:
        with open(OUT_PATH) as f:
            existing = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        existing = None

    data = resolve_gudi_village_names(data, existing)

    issues = find_unreferenced(data)
    if issues:
        print('VALIDATION ISSUES:')
        for issue in issues:
            print(f'  * {issue}')
        print()
        if validate_only:
            print(f'{len(issues)} issue(s) found. Fix the Excel and re-run.')
            sys.exit(1)
        print(f'{len(issues)} issue(s) found - proceeding anyway.')
        print()

    if validate_only:
        print('✓ Validation passed (no issues).')
        return

    narrative_sources = existing.get('narrative_sources', []) if existing else []

    output = {
        'metadata': build_metadata(),
        'phratries': build_phratries_list(data),
        'clans': data['clans'],
        'pens': data['pens'],
        'villages': data['villages'],
        'relationships': data['relationships'],
        'narrative_sources': narrative_sources,
    }

    for pen in output['pens']:
        if pen.get('gudi_village_id') is None:
            del pen['gudi_village_id']
        if not pen.get('village_ids'):
            del pen['village_ids']
        elif len(pen['village_ids']) == 0:
            del pen['village_ids']
        if not pen.get('villages'):
            del pen['villages']
        elif len(pen['villages']) == 0:
            del pen['villages']

    seen_rels = set()
    unique_rels = []
    for rel in output['relationships']:
        key = (rel['type'], rel['from_pen_id'], rel['to_pen_id'], rel['details'][:50])
        if key not in seen_rels:
            seen_rels.add(key)
            unique_rels.append(rel)
    output['relationships'] = unique_rels

    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f'✓ Wrote {OUT_PATH}')
    print(f'  Phratries: {len(output["phratries"])}')
    print(f'  Clans: {len(output["clans"])}')
    print(f'  Pens: {len(output["pens"])}')
    print(f'  Villages: {len(output["villages"])}')
    print(f'  Relationships: {len(output["relationships"])}')

    if issues:
        print()
        print(f'Note: {len(issues)} validation issue(s) were present (listed above).')


if __name__ == '__main__':
    main()
