"""Build human-friendly simple Excel workbook for analysts.

Output: data/gods_and_goddesses/Clan Gods Data - Simple.xlsx

Two sheets:
  Pens       — one row per pen, names-based relationships, no IDs required
  Data Dictionary — field descriptions, examples, and rules

Usage:
    python3 scripts/build_simple_excel.py
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DATA_PATH = 'data/gods_and_goddesses/clan_gods.json'
OUT_PATH = 'data/gods_and_goddesses/Clan Gods Data - Simple.xlsx'

with open(DATA_PATH) as f:
    cg = json.load(f)

pen_map = {p['id']: p for p in cg['pens']}

wb = Workbook()

header_font = Font(bold=True, color='ffffff', size=11)
header_fill = PatternFill('solid', fgColor='2a6f97')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc'),
)
readonly_fill = PatternFill('solid', fgColor='f0f0f0')
readonly_font = Font(color='888888', italic=True)

PEN_COLS = [
    'name', 'aliases', 'gender', 'clan', 'phratry',
    'village', 'village_details', 'palli', 'perma', 'karsad',
    'type', 'spouse', 'siblings', 'children', 'parent', 'notes'
]

CLAN_NAMES = [
    'Kunjam', 'Markami', 'Kalmu/Karma', 'Bhogam/Chote Telam',
    'Chote Telam/Bhogam', 'Bade Telam', 'Icham', 'Punem',
    'Kadiyam', 'Midiyam', 'Undam', 'Tati', 'Kadti', 'Rengo',
    'Oyam', 'Hemla', 'Tamo', 'Padami', 'Ujji/Dodi', 'Barse',
    'Madvi', 'Kawasi'
]

PHRATRY_NAMES = ['Kuhrami/Kadiari', 'Markami Kutumb', 'Madvi', 'Kawasi']

def make_id(name):
    import re
    n = name.lower().strip()
    n = re.sub(r'[^a-z0-9_\s]', '', n)
    n = re.sub(r'\s+', '_', n)
    return n

def resolve_pen_names(pen_ids):
    """Resolve pen IDs to display names."""
    names = []
    for pid in pen_ids:
        p = pen_map.get(pid)
        names.append(p['name'] if p else pid)
    return ', '.join(names)

def resolve_single_pen_name(pen_id):
    if not pen_id:
        return ''
    p = pen_map.get(pen_id)
    return p['name'] if p else pen_id

def clan_id_to_name(clan_id):
    if not clan_id:
        return ''
    for c in cg['clans']:
        if c['id'] == clan_id:
            return c['name']
    return clan_id

def phratry_id_to_name(phratry_id):
    if not phratry_id:
        return ''
    for p in cg['phratries']:
        if p['id'] == phratry_id:
            return p['name']
    return phratry_id

# ── Sheet 1: Pens ──
ws = wb.active
ws.title = 'Pens'
ws.sheet_properties.tabColor = '27ae60'

style_header = ws.cell
for i, col_name in enumerate(PEN_COLS, 1):
    c = ws.cell(row=1, column=i, value=col_name)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = thin_border

# Dropdowns
gender_dv = DataValidation('list', formula1='"male,female,unknown"', allow_blank=True)
ws.add_data_validation(gender_dv)

clan_dv = DataValidation('list',
    formula1='"%s"' % ','.join(CLAN_NAMES), allow_blank=True)
ws.add_data_validation(clan_dv)

phratry_dv = DataValidation('list',
    formula1='"%s"' % ','.join(PHRATRY_NAMES), allow_blank=True)
ws.add_data_validation(phratry_dv)

type_dv = DataValidation('list', formula1='"main,subordinate"', allow_blank=True)
ws.add_data_validation(type_dv)

# Build cross-reference: pen name -> list of related pen names
def get_related_names(pen_id, rel_type, direction='from'):
    names = []
    for r in cg['relationships']:
        if direction == 'from' and r[direction + '_pen_id'] == pen_id and r['type'] == rel_type:
            target_id = r['to_pen_id']
            p = pen_map.get(target_id)
            if p:
                names.append(p['name'])
        elif direction == 'to' and r[direction + '_pen_id'] == pen_id and r['type'] == rel_type:
            target_id = r['from_pen_id']
            p = pen_map.get(target_id)
            if p:
                names.append(p['name'])
    return names

# Write data rows
for r, pen in enumerate(cg['pens'], 2):
    pen_id = pen['id']
    
    # Resolve relationships
    spouses = []
    # spouse: marriage where this pen is the 'from' pen (marriage to)
    for rel in cg['relationships']:
        if rel['type'] == 'marriage' and rel['from_pen_id'] == pen_id:
            p = pen_map.get(rel['to_pen_id'])
            if p:
                spouses.append(p['name'])
        # Also check where this pen is the 'to' pen of a marriage
        if rel['type'] == 'marriage' and rel['to_pen_id'] == pen_id:
            p = pen_map.get(rel['from_pen_id'])
            if p and p['name'] not in spouses:
                spouses.append(p['name'])
        # ghar_jamai: spouse is the one they're ghar jamai to
        if rel['type'] == 'ghar_jamai' and rel['from_pen_id'] == pen_id:
            p = pen_map.get(rel['to_pen_id'])
            if p:
                spouses.append(f'ghar_jamai:{p["name"]}')
    
    siblings = get_related_names(pen_id, 'sibling', 'from')
    # Also get siblings from 'to' direction
    for rel in cg['relationships']:
        if rel['type'] == 'sibling' and rel['to_pen_id'] == pen_id and rel['from_pen_id'] != pen_id:
            p = pen_map.get(rel['from_pen_id'])
            if p and p['name'] not in siblings:
                siblings.append(p['name'])
    
    children = get_related_names(pen_id, 'parent', 'from')
    parent = get_related_names(pen_id, 'parent', 'to')
    
    vals = [
        pen['name'],
        ', '.join(pen.get('aliases', [])),
        pen.get('gender', 'unknown'),
        clan_id_to_name(pen.get('clan_id')),
        phratry_id_to_name(pen.get('phratry_id')),
        pen.get('gudi_village') or '',
        '',
        ', '.join(pen.get('palli', [])),
        pen.get('perma', ''),
        pen.get('karsad', ''),
        pen.get('type', 'main'),
        ', '.join(spouses),
        ', '.join(siblings),
        ', '.join(children),
        resolve_single_pen_name(parent[0]) if parent else '',
        pen.get('notes', ''),
    ]
    
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = cell_align
        cell.border = thin_border
        if c == 1:  # name column - required
            cell.font = Font(bold=True)
    
    # Add dropdowns
    gender_cell = ws.cell(row=r, column=3)
    gender_dv.add(gender_cell)
    clan_cell = ws.cell(row=r, column=4)
    clan_dv.add(clan_cell)
    phratry_cell = ws.cell(row=r, column=5)
    phratry_dv.add(phratry_cell)
    type_cell = ws.cell(row=r, column=11)
    type_dv.add(type_cell)

# Column widths
col_widths = {
    1: 22, 2: 30, 3: 10, 4: 20, 5: 20,
    6: 25, 7: 35, 8: 35, 9: 20, 10: 20,
    11: 12, 12: 25, 13: 30, 14: 30, 15: 22, 16: 40
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:P{len(cg["pens"]) + 1}'

# ── Sheet 2: Data Dictionary ──
ws2 = wb.create_sheet('Data Dictionary', 1)
ws2.sheet_properties.tabColor = '3498db'

dict_cols = ['Field', 'Description', 'Example', 'Required', 'Rules']
dict_header_fill = PatternFill('solid', fgColor='3498db')

for i, name in enumerate(dict_cols, 1):
    c = ws2.cell(row=1, column=i, value=name)
    c.font = Font(bold=True, color='ffffff', size=11)
    c.fill = dict_header_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

dictionary = [
    ['name', 'Name of the pen (deity/god). Must be unique.', 'Godar Hunga', 'Yes', 'Will be used to auto-generate a pen ID. Duplicate names will be flagged by the converter.'],
    ['aliases', 'Alternate names or spellings for this pen.', 'Godar Hunga, Godar Hunga', 'No', 'Comma-separated. Used for lookup when resolving relationships.'],
    ['gender', 'Gender of the pen.', 'male', 'Yes', 'Dropdown: male / female / unknown. Female indicators in name auto-detect (dokri, lungo, etc.).'],
    ['clan', 'Clan this pen belongs to.', 'Bade Telam', 'Yes if known', 'Dropdown from 22 known clans. If unknown, leave blank.'],
    ['phratry', 'Phratry this pen belongs to.', 'Markami Kutumb', 'Yes if known', 'Dropdown: Kuhrami/Kadiari, Markami Kutumb, Madvi, Kawasi.'],
    ['village', 'Gudi village — where the pen\'s shrine is located.', 'Jhirka', 'No', 'Free text. Same pen name + village across different phratries = same village is reused; if names match but phratries differ, they\'re treated as separate village entries.'],
    ['village_details', 'Extra info about the gudi village: location, landmarks, Google Maps link, nearby villages, lat/long.', 'Near Jhirka waterfall, below the hill. 3 km from Kamalnar.', 'No', 'Free text. Not used directly on the map — kept for analyst reference.'],
    ['palli', 'Territory of influence — villages/areas where this pen is worshipped.', 'Dokometta, Dhanora, Tumnar', 'No', 'Comma-separated village names. These are NOT the gudi village — they are additional villages that bring offerings here.'],
    ['perma', 'Name of the priest (perma) who performs rituals at this pen\'s shrine.', 'Joga Telam', 'No', 'Free text.'],
    ['karsad', 'Annual festival day when offerings are made to this pen.', 'Tuesday in February', 'No', 'Free text. Examples: "Tuesday in February", "Diwad and Bijja Pandum".'],
    ['type', 'Whether this is the main pen of its village or a subordinate pen.', 'main', 'Yes', 'Dropdown: main / subordinate.'],
    ['spouse', 'Spouse(s) of this pen. Use comma for multiple.', 'Punga Rungi', 'No', 'Pen name(s). Prefix with "ghar_jamai:" for ghar jamai (e.g. "ghar_jamai:Punga Rungi"). The converter will parse these into relationship entries.'],
    ['siblings', 'Sibling pens (brothers/sisters).', 'Huru Mara, Daro Moitor, Bhimaraj, Godar Bhima', 'No', 'Comma-separated pen names. The converter creates sibling relationship pairs.'],
    ['children', 'Children of this pen.', 'Lug Unga, Punga Rungi', 'No', 'Comma-separated pen names. The converter creates parent-child relationship entries.'],
    ['parent', 'Parent (father/mother) of this pen.', 'Huru Mara', 'No', 'Single pen name. The converter creates a child-parent relationship entry.'],
    ['notes', 'Any additional notes about this pen.', 'Son of Huru Mara and Katta Bodke (3rd wife)', 'No', 'Free text.'],
]

for r, row_data in enumerate(dictionary, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border
        if c == 1:
            cell.font = Font(bold=True)
        if c == 3 or c == 5:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 55
ws2.column_dimensions['C'].width = 45
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 55
ws2.freeze_panes = 'A2'

wb.save(OUT_PATH)
print(f"✓ Wrote {OUT_PATH}")
print(f"  Pens: {len(cg['pens'])} rows pre-populated")
print(f"  Data Dictionary: {len(dictionary)} field descriptions")
