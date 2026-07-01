import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment

SIMPLE_XLSX = 'data/gods_and_goddesses/Clan Gods Data - Simple.xlsx'
MAIN_XLSX = 'data/gods_and_goddesses/Clan Gods Data.xlsx'
BHUVAN_GEOJSON = 'data/bhuvan_villages_merged.geojson'
CENTROIDS_JSON = 'data/gods_and_goddesses/village_centroids.json'

# ── 1. Build village → v_code lookup ──

lookup = {}  # village_name_lower -> v_code

# Source A: main Excel
wb_main = load_workbook(MAIN_XLSX, data_only=True)
ws_villages = wb_main['Villages']
for row in ws_villages.iter_rows(min_row=2, max_row=ws_villages.max_row, values_only=True):
    name = row[0]
    v_code = row[9] if len(row) > 9 else None
    if name and v_code:
        try:
            v_code = str(int(float(v_code)))
        except (ValueError, TypeError):
            continue
        key = name.strip().lower()
        # Don't overwrite if already exists (first wins)
        if key not in lookup:
            lookup[key] = v_code
print(f"Main Excel: {len(lookup)} villages")

# Source B: village_centroids.json
if Path(CENTROIDS_JSON).exists():
    with open(CENTROIDS_JSON) as f:
        centroids = json.load(f)
    for vname, vdata in centroids.items():
        code = vdata.get('code')
        if code:
            key = vname.strip().lower()
            if key not in lookup:
                lookup[key] = str(code)
print(f"After centroids: {len(lookup)} villages")

# Source C: Bhuvan GeoJSON — match via NAME_OVERRIDES + prefix matching
NAME_OVERRIDES = {
    'Itawar': 'Hitawar',
    'Pharaspal': 'Faraspal',
    'Kesapur': 'Keshapur',
    'Katural, Pumbad (near Gangalur)': 'Katur',
    'Kamalnar? Kamalur': 'Kamaloor',
    'Manganar': 'Mangalnar',
    'Rekavaya': 'Rekavaya',
    'Vechapal': 'Vechapal',
    'Vengpal (Vengur)': 'Vengpal',
    'Vengur': 'Vengur',
    'Benpal/Bayampal': 'Bengpal',
    'Benpal': 'Bengpal',
    'Omalwar/Samalwar': 'Samalwar',
    'Omalwar': 'Samalwar',
    'Mirtulnar/Midkulnar': 'Mirtulnar',
}

PREFERRED_DISTRICTS = ['Dakshin Bastar Dantewada', 'Dantewada', 'Bijapur', 'Bastar', 'Sukma', 'Kondagaon']

with open(BHUVAN_GEOJSON) as f:
    bhuvan = json.load(f)

# Index Bhuvan features by name
name_to_features = {}
for feat in bhuvan['features']:
    props = feat.get('properties', {})
    vname = (props.get('v_name') or '').strip()
    vcode = str(props.get('v_code') or '')
    dname = (props.get('d_name') or '').strip()
    if not vname or not vcode:
        continue
    entry = {'v_name': vname, 'v_code': vcode, 'd_name': dname, 'feature': feat}
    name_to_features.setdefault(vname.lower(), []).append(entry)

def find_features_by_prefix(keyword):
    keyword = keyword.lower().strip()
    results = []
    for bname, entries in name_to_features.items():
        if bname == keyword:
            results.extend(entries)
    for bname, entries in name_to_features.items():
        if bname != keyword and (bname.startswith(keyword + ' ') or bname.startswith(keyword + '(')):
            results.extend(entries)
    return results

def pick_best_feature(entries):
    if not entries:
        return None
    if len(entries) == 1:
        return entries[0]
    for dist in PREFERRED_DISTRICTS:
        for e in entries:
            if e['d_name'] == dist:
                return e
    return entries[0]

def extract_base_names(name):
    import re
    name = re.sub(r'\(.*?\)', '', name).strip()
    parts = re.split(r'[/,;?]', name)
    bases = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        words = part.split()
        if len(words) >= 1 and len(words[0]) >= 3:
            bases.append(words[0])
    return bases

def match_to_bhuvan(clan_gods_name):
    name_lower = clan_gods_name.strip().lower()
    if name_lower in lookup:
        return lookup[name_lower]
    override = NAME_OVERRIDES.get(clan_gods_name.strip())
    if override:
        entries = name_to_features.get(override.lower(), [])
        best = pick_best_feature(entries)
        if best:
            return best['v_code']
    bases = extract_base_names(clan_gods_name)
    for base in bases:
        entries = find_features_by_prefix(base)
        best = pick_best_feature(entries)
        if best:
            return best['v_code']
    return None

# Try matching remaining villages
ALL_SIMPLE_VILLAGES = [
    'Benpal/Bayampal', 'Bhansi', 'Bododi', 'Cholnar', 'Daler (near Marh)',
    'Dhurli', 'Dugeli', 'Etlapad', 'Gongpal', 'Gumiyapal (Guyempad)',
    'Hiroli', 'Jawanga', 'Jhadka (Orcha Block)', 'Jhirka', 'Kamalnar/Kamalur',
    'Kamkajojor', 'Kamkanar', 'Kesapur', 'Kodoli', 'Kondapal',
    'Kunjampara, near Ganjenar', 'Kuwe', 'Madadi', 'Madpal', 'Manganar',
    'Markagudem', 'Metapal', 'Mirtulnar/Midkulnar', 'Omalwar/Samalwar',
    'Palnar', 'Pandewar', 'Pharaspal', 'Pidiya', 'Pinkonda', 'Pondum',
    'Rekavaya', 'Renganar', 'Shyamgiri', 'Tadopadar', 'Tamirguda',
    'Tamodi', 'Tikanpal', 'Vechapal', 'Vengpal (Vengur)', 'Vengur',
]

for vname in ALL_SIMPLE_VILLAGES:
    key = vname.strip().lower()
    if key in lookup:
        continue
    code = match_to_bhuvan(vname)
    if code:
        lookup[key] = code
        print(f"  Matched: {vname:35s} → {code}")
    else:
        print(f"  NO MATCH: {vname:35s}")

print(f"\nTotal lookup entries: {len(lookup)}")

# ── 2. Update Simple Excel ──
wb = load_workbook(SIMPLE_XLSX)
ws = wb['Pens']

updated = 0
skipped_existing = 0
no_match = 0

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=16):
    vals = [c.value for c in row]
    pen_name = str(vals[0] or '')
    if '═══' in pen_name:
        continue
    village = str(vals[5] or '').strip()
    existing_details = str(vals[6] or '').strip()
    if not village:
        continue
    if existing_details:
        skipped_existing += 1
        continue  # user's manual entry
    v_code = lookup.get(village.lower())
    if v_code:
        row[6].value = v_code
        row[6].alignment = Alignment(wrap_text=True, vertical='top')
        updated += 1
        print(f"  Filled: {pen_name:40s} @ {village:30s} → {v_code}")
    else:
        no_match += 1

wb.save(SIMPLE_XLSX)
print(f"\nDone: {updated} filled, {skipped_existing} already had codes, {no_match} no match")
